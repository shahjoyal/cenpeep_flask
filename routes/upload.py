"""
upload.py — Flask upload route for CENPEEP
==========================================
Accepts .xlsx / .xls files and does two things:

1. SMART MULTI-SHEET PARSE (new):
   • Reads every sheet in the workbook.
   • For each sheet, tries to identify CENPEEP input fields by scanning:
       – the standard "CenPeep Corrected" column layout (col A=particulars,
         col B=UOM, col C=symbol, col D=formula/INPUT, col E=value)
       – a "raw / field-name header" layout where the first row contains
         field names and subsequent rows are data rows (handles multiple
         load readings → averages them).
   • Returns per-sheet metadata + a merged `extracted` dict (last-wins
     for conflicts; the CenPeep sheet always wins if present).

2. LEGACY SINGLE-SHEET PARSE (kept for backward compat).
"""

"""
upload.py — Flask upload route for CENPEEP  (v2: ML detection + chunked parsing)
==================================================================================
Accepts .xlsx / .xls files and does three things:

1. SMART MULTI-SHEET PARSE — reads every sheet, tries the strict CenPeep
   column layout first, then a generic raw-tabular layout.

2. ML FIELD DETECTION (new) — for any column NOT matched by the rule-based
   symbol/label lookup, a basic trainable TF-IDF + cosine-similarity model
   (see ml/field_classifier.py) scores the header text against known
   CENPEEP field phrasings and assigns a field id if confident enough.
   This is what lets a sheet with totally different header wording (e.g.
   "MAIN STM TEMP-L", "Primary Air APH Temp I/L A") still get its columns
   identified, instead of relying only on exact alias matches.

3. CHUNKED / STREAMED PARSING (new) — large sheets (many rows and/or wide
   column counts) are read via openpyxl's read_only streaming mode and
   processed in fixed-size row chunks, so we never hold the full sheet plus
   multiple copies of it in memory at once. This is what allows bigger
   files (the route's max upload size has also been raised) to be parsed
   without timing out or exhausting memory.

Multiple load/data readings for the same field → averaged automatically
(unchanged from v1, now chunk-aware).
"""

import io
import os
import re
import time
import datetime
import statistics
from flask import Blueprint, request, jsonify

try:
    from dateutil import parser as _dateutil_parser
    HAS_DATEUTIL = True
except ImportError:
    HAS_DATEUTIL = False

try:
    import python_calamine
    HAS_CALAMINE = True
except ImportError:
    HAS_CALAMINE = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

from ml.field_classifier import get_classifier, DEFAULT_CONFIDENCE_THRESHOLD
from ml.training_data import is_non_field_header

upload_bp = Blueprint('upload', __name__)

# ─── Chunking config ───────────────────────────────────────────────────────────
CHUNK_ROWS = 300          # rows processed per chunk for large sheets
LARGE_SHEET_ROW_THRESHOLD = 500   # sheets bigger than this use chunked streaming
# How many leading rows we scan to find the header row. Real plant exports
# often stack a report title / merged-cell banner row (and sometimes a
# blank spacer row above that) before the actual column-header row -- e.g.
# 4 blank rows, then a "<Sheet> Report" title row, then the header row on
# the 6th physical row. 5 was tight enough that a sheet with just one title
# row on top of the usual blank padding pushed its header row one row past
# the scan window, so the header was never even considered as a candidate
# and the whole sheet fell through to "unrecognized" (see the Coal & Ash
# Analysis sheet case: 4 blank rows + 1 title row put the header at row
# index 5, outside rows[:5]). Widened to 10 to comfortably cover that
# pattern; _find_header_row still only keeps rows that look header-shaped,
# so the wider window doesn't make it more likely to pick a wrong row.
HEADER_SCAN_ROWS = 10

# ─── Symbol → CENPEEP field-id map ────────────────────────────────────────────
SYM_MAP = {
    'L': 'L', 'Ffw': 'Ffw', 'Fin': 'Fin',
    'Cba': 'Cba', 'Cfa': 'Cfa', 'Pfa': 'Pfa', 'Pba': 'Pba',
    'M': 'M', 'A': 'A', 'VM': 'VM', 'FC': 'FC', 'GCV': 'GCV', 'S': 'S',
    'O2in': 'O2in', 'COin': 'COin', 'O2out': 'O2out', 'COout': 'COout',
    'Tgi': 'Tgi', 'Tgo': 'Tgo',
    'Tpai': 'Tpai', 'Tpao': 'Tpao', 'Tsai': 'Tsai', 'Tsao': 'Tsao',
    'Fsa': 'Fsa', 'Fpa': 'Fpa', 'Tref': 'Tref',
    # Design — proximate
    'Md': 'Md', 'Ad': 'Ad', 'VMd': 'VMd', 'FCd': 'FCd',
    # Design — ultimate
    'Cd': 'Cd', 'Sd': 'Sd', 'Hd': 'Hd', 'Nd': 'Nd', 'Od': 'Od',
    'Gcvd': 'GCVd', 'GCVd': 'GCVd', 'Trad': 'Trad', 'Mwvd': 'Mwvd',
    # BEE (BEE-2 Indirect / heat-loss method, public/tab3.html +
    # script3.js) field ids that don't already exist as a CENPEEP
    # symbol. M/A/GCV/Cba/Cfa are the SAME physical quantities BEE uses
    # (coal moisture/ash/GCV, unburnt carbon in bottom/fly ash) and
    # already resolve via the entries above, so they're intentionally
    # not repeated here.
    'O2fg': 'O2fg', 'COfg': 'COfg', 'CO2fg': 'CO2fg', 'Tfg': 'Tfg',
    'Tamb': 'Tamb', 'Hum': 'Hum',
    'C': 'C', 'H2': 'H2', 'N2': 'N2', 'O2f': 'O2f',
    'GCVba': 'GCVba', 'GCVfa': 'GCVfa',
    'BL': 'BL', 'SP': 'SP',
}

# Also accept case-insensitive & common variants
SYM_MAP_LOWER = {k.lower(): v for k, v in SYM_MAP.items()}

# Fields that are ALWAYS manual entry — never auto-detected from an upload,
# by any strategy (CenPeep-layout symbol match, rule-based header alias, or
# ML fallback), and never flagged as "missing" or colored red/green. Pfa/Pba
# ("% of Fly/Bottom Ash in Total Ash") are plant-specific split ratios that
# don't reliably show up as their own column on real sheets — treating them
# as always-manual avoids both false "detected" guesses and noisy "missing"
# flags for a field that was never expected to be found anyway.
# Sd/GCVd/Trad/Mwvd (Design Conditions — Ultimate Analysis: Sulfur, GCV,
# Ref. Air Temp, Moisture in Air) have no proximate-side equivalent to
# derive from — same as As-Fired Sulfur/GCV, they're always typed in by
# hand — so they get the same always-manual treatment here.
# Md/Ad/VMd/FCd (Design — Proximate) carry the same "MANUAL" tag on the
# calculator form and the same "never auto-filled from an upload" intent
# (see the comment above the Design — Proximate section in
# public/calculator.html) — added here so that intent is actually
# enforced, instead of only being true in the UI copy. Previously these
# could still be picked up from the strict CenPeep column layout (a sheet
# listing the symbol "Md"/"Ad" twice — once as-fired, once design) via the
# duplicate-symbol handling in _parse_cenpeep_layout(); that handling still
# runs (so the second occurrence still correctly resolves to the AUTO/
# readonly Md2/Ad2 slot for the summary table), but the first occurrence
# no longer auto-fills or colors the manual Md/Ad inputs.
# S (As-Fired Sulfur), COin (Avg. Flue Gas CO — APH In), and Tref (Design
# Ambient / Ref Air Temp) are also now always-manual — same reasoning as
# Sd/GCVd/Trad/Mwvd above: these are typed in by hand from a lab report or
# a fixed plant design value rather than reliably reported as their own
# DCS/log column, so they get the same "never auto-detected, never
# colored detected/missing" treatment. Mirrored on the calculator form with
# the "MANUAL" tag (see public/calculator.html) and removed from
# REQUIRED_FIELDS below, same as every other NEVER_AUTO_DETECT field.
NEVER_AUTO_DETECT = {
    'Pfa', 'Pba', 'Sd', 'GCVd', 'Trad', 'Mwvd', 'Md', 'Ad', 'VMd', 'FCd',
    'S', 'COin', 'Tref',
    # L6 (BEE-2 Indirect — Radiation & Unaccounted Losses) carries the
    # same "MANUAL" tag on public/tab3.html that every other field in
    # this set carries on its own calculator form — it's an assumed/
    # judgement figure (typically 1-2%), never a column a plant sheet
    # reports, so it gets the same never-auto-detect / never-flagged-
    # missing treatment as the rest of this set.
    'L6',
}

# ─── Full list of CENPEEP input fields the calculator needs ──────────────────
# This is every editable (non-auto-computed) field on the calculator form —
# used only to report, after a file is parsed, which required fields were
# NOT found on the selected sheet (CO2in/CO2out and Cd/Hd/Md2/Nd/Od/Ad2 are
# excluded: those are AUTO/readonly fields computed by the app — the latter
# client-side from Md/Ad/VMd/FCd — never read from an upload).
# Pfa/Pba/Sd/GCVd/Trad/Mwvd/Md/Ad/VMd/FCd are deliberately excluded — see
# NEVER_AUTO_DETECT above. Being always-manual, they should never show up
# as either "detected" (green) or "missing" (red) — in the upload summary,
# the field coloring on the calculator form, or the r.py Word report.
REQUIRED_FIELDS = [
    'L', 'Ffw', 'Fin', 'Cba', 'Cfa',
    'M', 'A', 'VM', 'FC', 'GCV',
    'O2in', 'O2out', 'COout',
    'Tgi', 'Tgo', 'Tpai', 'Tpao', 'Tsai', 'Tsao', 'Fsa', 'Fpa',
]

# Full human-readable name for each field id, taken verbatim from the
# "field-name" label next to each input on the calculator form (public/
# calculator.html) — used anywhere a field is shown to a person (reports,
# the upload summary) instead of the bare symbol/abbreviation.
FIELD_LABELS = {
    'L': 'Unit Load', 'Ffw': 'Steam Flow', 'Fin': 'Total Coal Flow',
    'Cba': 'Unburnt Carbon in Bottom Ash', 'Cfa': 'Unburnt Carbon in Fly Ash',
    'Pfa': '% of Fly Ash in Total Ash', 'Pba': '% of Bottom Ash in Total Ash',
    'M': 'Moisture', 'A': 'Ash', 'VM': 'Volatile Matter', 'FC': 'Fixed Carbon',
    'GCV': 'Gross Calorific Value (GCV)', 'S': 'Sulfur',
    'O2in': 'Avg. Flue Gas O\u2082 \u2014 APH In', 'COin': 'Avg. Flue Gas CO \u2014 APH In',
    'O2out': 'Avg. Flue Gas O\u2082 \u2014 APH Out', 'COout': 'Avg. Flue Gas CO \u2014 APH Out',
    'Tgi': 'Avg. Flue Gas Temp \u2014 APH In', 'Tgo': 'Avg. Flue Gas Temp \u2014 APH Out',
    'Tpai': 'Primary Air to APH Temp In', 'Tpao': 'Primary Air from APH Temp Out',
    'Tsai': 'Secondary Air to APH Temp In', 'Tsao': 'Secondary Air from APH Temp Out',
    'Fsa': 'Total Secondary Air Flow', 'Fpa': 'Total Primary Air Flow',
    'Tref': 'Design Ambient / Ref Air Temp',
    'Md': 'Moisture \u2014 Design', 'Ad': 'Ash \u2014 Design',
    'VMd': 'Volatile Matter \u2014 Design', 'FCd': 'Fixed Carbon \u2014 Design',
    'Cd': 'Carbon \u2014 Design', 'Sd': 'Sulfur \u2014 Design', 'Hd': 'Hydrogen \u2014 Design',
    'Md2': 'Moisture \u2014 Design (Ultimate)', 'Nd': 'Nitrogen \u2014 Design',
    'Od': 'Oxygen \u2014 Design', 'Ad2': 'Ash \u2014 Design (Ultimate)',
    'GCVd': 'GCV \u2014 Design', 'Trad': 'Ref. Air Temp \u2014 Design',
    'Mwvd': 'Moisture in Air \u2014 Design',
    # BEE (BEE-2 Indirect) labels, taken verbatim from public/tab3.html /
    # script3.js's INPUT_LABELS. M/A/GCV/Cba/Cfa already have labels above
    # and aren't repeated.
    'O2fg': 'O2 in Flue Gas', 'COfg': 'CO in Flue Gas', 'CO2fg': 'CO2 in Flue Gas',
    'Tfg': 'Avg. Flue Gas Temperature', 'Tamb': 'Ambient Temperature',
    'Hum': 'Humidity in Ambient Air',
    'C': 'Carbon', 'H2': 'Hydrogen', 'N2': 'Nitrogen', 'O2f': 'Oxygen',
    'GCVba': 'GCV of Bottom Ash', 'GCVfa': 'GCV of Fly Ash',
    'BL': 'Boiler Load', 'SP': 'Steam Pressure',
}

# Human-readable label guesses for unknown-layout headers
LABEL_ALIASES = {
    'load': 'L', 'unit load': 'L', 'mw': 'L',
    'steam flow': 'Ffw', 'steamflow': 'Ffw',
    'coal flow': 'Fin', 'coalflow': 'Fin', 'fuel flow': 'Fin',
    'moisture': 'M', 'ash': 'A',
    'volatile matter': 'VM', 'vm': 'VM',
    'fixed carbon': 'FC', 'fc': 'FC',
    'gcv': 'GCV', 'gross calorific value': 'GCV',
    'sulphur': 'S', 'sulfur': 'S',
    'o2 in': 'O2in', 'o2in': 'O2in',
    'o2 out': 'O2out', 'o2out': 'O2out',
    'co in': 'COin', 'coin': 'COin',
    'co out': 'COout', 'coout': 'COout',
    'fg temp in': 'Tgi', 'flue gas temp in': 'Tgi',
    'fg temp out': 'Tgo', 'flue gas temp out': 'Tgo',
    'pa temp in': 'Tpai', 'sa temp in': 'Tsai',
    'pa temp out': 'Tpao', 'sa temp out': 'Tsao',
    'pa flow': 'Fpa', 'sa flow': 'Fsa',
    # Exact real-sheet DCS tag: "TEMP AT APH O/L MEDIAN" — on its own this
    # phrasing is ambiguous with Tgo (Flue Gas Temp APH Out also gets
    # called "APH O/L ... Temp"), so it's handled as a hard-coded exact
    # alias rather than left to the ML classifier's fuzzy similarity,
    # which was pulling it toward Tgo. Confirmed against the actual
    # workbook this represents Primary Air from APH Temp Out (Tpao).
    'temp at aph ol median': 'Tpao',
    # Exact real-sheet DCS tags: "OXYGEN IN FLUE GAS(L)-AH O/L" and
    # "OXYGEN IN FLUE GAS(R)-AH O/L" — Left/Right duct O2 sensors at the
    # APH OUTLET (the trailing "AH O/L" = "Air Heater Outlet"). The ML
    # classifier was matching these to O2in instead (score ~0.63, above
    # threshold) because the word "IN" in "OXYGEN IN FLUE GAS" reads as
    # "inlet" out of context, when it's actually just "oxygen [found] in
    # [the] flue gas" — the real inlet/outlet qualifier is the "AH O/L"
    # suffix, which means outlet. Hard-coded here as exact aliases (rather
    # than left to the fuzzy ML match) so both L/R columns resolve to
    # O2out deterministically and both get picked up by
    # MULTI_COLUMN_AVERAGE_FIELDS / _dedupe_columns_per_field, which
    # averages the two sensor readings together instead of keeping only
    # one.
    'oxygen in flue gaslah ol': 'O2out',
    'oxygen in flue gasrah ol': 'O2out',
    'unburnt bottom': 'Cba', 'unburnt fly': 'Cfa',
    'unburnts in bottom ash': 'Cba', 'unburnts in fly ash': 'Cfa',
    'unburnt in bottom ash': 'Cba', 'unburnt in fly ash': 'Cfa',
    'bottom ash unburnts': 'Cba', 'fly ash unburnts': 'Cfa',
    # NOTE: bare "Bottom Ash"/"Fly Ash" (%) used to be hard-mapped straight
    # to Pba/Pfa ("% of ash in total ash") right here. That was wrong for
    # real lab-report / LOI / boiler-efficiency sheets, where a bare
    # "Bottom Ash (%)" or "Fly Ash %" column is actually the loss-on-
    # ignition / unburnt-carbon test result (Cba/Cfa), not the ash-split
    # percentage — and because this was an exact-string RULE match it ran
    # before ML ever got a chance, so it silently stole the column and
    # left Cba/Cfa undetected every single time. Disambiguation between
    # the two now lives in _match_tag_patterns() below (keyed on whether
    # "total" appears in the header), since both phrasings otherwise look
    # identical.

    # ── BEE (BEE-2 Indirect) exact-label aliases ──────────────────────────
    # Taken verbatim from public/tab3.html's own field-name text (plus a
    # couple of obvious variants) so a sheet that echoes the calculator's
    # own wording — or a common lab-report phrasing — resolves via the
    # fast exact-match path instead of relying solely on the ML fallback.
    # Kept narrowly scoped (no bare "oxygen"/"o2" collisions with the
    # APH-in/out O2in/O2out family above, which always require an in/out
    # qualifier) so nothing here can steal a column from an existing
    # CENPEEP alias.
    'o2 in flue gas': 'O2fg', 'oxygen in flue gas': 'O2fg',
    'o2 flue gas': 'O2fg', 'flue gas o2': 'O2fg', 'o2fg': 'O2fg',
    'co in flue gas': 'COfg', 'carbon monoxide in flue gas': 'COfg',
    'co flue gas': 'COfg', 'flue gas co': 'COfg', 'cofg': 'COfg',
    'co2 in flue gas': 'CO2fg', 'carbon dioxide in flue gas': 'CO2fg',
    'co2 flue gas': 'CO2fg', 'flue gas co2': 'CO2fg', 'co2fg': 'CO2fg',
    'average flue gas temperature': 'Tfg', 'avg flue gas temperature': 'Tfg',
    'flue gas temperature': 'Tfg', 'flue gas temp': 'Tfg', 'fg temperature': 'Tfg',
    'ambient temperature': 'Tamb', 'atmospheric temperature': 'Tamb',
    'ambient temp': 'Tamb', 'atmospheric air temperature': 'Tamb',
    'humidity in ambient air': 'Hum', 'humidity': 'Hum',
    'ambient humidity': 'Hum', 'relative humidity': 'Hum',
    'carbon': 'C', 'hydrogen': 'H2', 'nitrogen': 'N2', 'oxygen': 'O2f',
    'ash content': 'A', 'gcv of coal': 'GCV',
    'gcv of bottom ash': 'GCVba', 'gcv bottom ash': 'GCVba',
    'gcv of fly ash': 'GCVfa', 'gcv fly ash': 'GCVfa',
    'boiler load': 'BL', 'steam pressure': 'SP',
}


# ─── Highlighted-column detection ─────────────────────────────────────────────
# Real plant sheets often have the engineer manually highlight (yellow-fill,
# usually) exactly the columns that matter for the CENPEEP efficiency calc,
# out of dozens/hundreds of DCS tag columns on the sheet. That's a strong
# human-provided signal we were previously throwing away entirely — only
# header TEXT was ever used to decide a column's field. Two failure modes
# this caused in practice:
#   1. A highlighted column with unusual/abbreviated wording scored just
#      under the ML confidence threshold and was silently skipped, while a
#      differently-worded but WRONG column elsewhere matched confidently.
#   2. When two columns (one highlighted, one not) both matched the same
#      field (e.g. "TM(ARB)" and "IM(ADB)" both loosely mean "moisture"),
#      their readings were averaged TOGETHER — quietly blending the
#      engineer's chosen reading with an unrelated one and corrupting the
#      value, with no visible sign anything had gone wrong.
# The functions below read (once per upload, cheaply — only the first few
# header rows, not the whole sheet) which header cells carry a real fill
# color, so that signal can be used to (a) prefer the highlighted column
# whenever it conflicts with a non-highlighted one mapped to the same
# field, and (b) retry unmatched highlighted columns at a lower confidence
# threshold before giving up on them.
HIGHLIGHT_SCAN_ROWS = HEADER_SCAN_ROWS  # only need the header-row candidates
# A relaxed threshold used ONLY for header cells we already know were
# deliberately highlighted by a human — still requires real similarity
# (this is not "accept anything"), just less margin than the default.
HIGHLIGHTED_ML_THRESHOLD = 0.30


def _is_highlighted_fill(cell):
    """
    True if a cell has a real (non-white/none) solid fill color.
    Guards every attribute access - openpyxl fill/color objects raise on
    some malformed theme-color entries (e.g. a cell with a theme tint but
    an unreadable rgb value, seen on real exports).
    """
    try:
        fill = cell.fill
        if fill is None or fill.patternType != 'solid':
            return False
        fg = fill.fgColor
        if fg is None:
            return False
        if getattr(fg, 'type', None) == 'rgb':
            rgb = fg.rgb
            if not isinstance(rgb, str):
                return False
            # Treat pure white / fully transparent as "not highlighted" -
            # a solid-white fill is usually just a formatting artifact,
            # not an intentional highlight.
            return rgb.upper() not in ('00000000', 'FFFFFFFF', 'FFFFFF')
        if getattr(fg, 'type', None) == 'theme':
            # A theme color WAS explicitly set on this cell (as opposed to
            # no fill at all) - treat that as "highlighted" too, since some
            # workbooks use a theme-based accent color instead of a raw RGB
            # yellow.
            return True
        return False
    except Exception:
        return False


def _scan_header_highlights(file_bytes):
    """
    Returns {sheetName: {rowIdx: {colIdx, ...}}} - for each sheet, the set
    of column indices that carry a highlighted fill, for each of the first
    HIGHLIGHT_SCAN_ROWS rows (0-indexed). Only openpyxl exposes cell style
    info (calamine reads values only, for speed), so this always does a
    second, SEPARATE, lightweight openpyxl read_only pass restricted to a
    handful of rows - cheap even on very large workbooks, since read_only
    iteration is lazy and we never touch a data row.
    Returns {} (feature silently disabled) if openpyxl isn't available or
    the file can't be opened a second time this way - highlight detection
    is a bonus signal, never a requirement for a successful parse.
    """
    if not HAS_OPENPYXL:
        return {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return {}

    result = {}
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            if not hasattr(ws, 'iter_rows'):
                continue
            sheet_map = {}
            try:
                for r_idx, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=HIGHLIGHT_SCAN_ROWS)
                ):
                    cols = {c_idx for c_idx, cell in enumerate(row) if _is_highlighted_fill(cell)}
                    if cols:
                        sheet_map[r_idx] = cols
            except Exception:
                # A malformed row/style shouldn't take down highlight
                # detection for the rest of the sheet or other sheets.
                pass
            if sheet_map:
                result[name] = sheet_map
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return result


def _apply_highlight_signal(col_map, col_source, col_confidence, headers,
                             highlighted_cols, use_ml=True):
    """
    Adjusts a column->field mapping using the set of highlighted column
    indices for this sheet's header row:

      1. Conflict resolution - if a field id is currently backed by more
         than one column and at least one (but not all) of them is
         highlighted, drop the non-highlighted column(s) for that field
         entirely. The engineer's marked column wins outright rather than
         being averaged together with a look-alike column.
      2. Recovery - any highlighted column that matched NO field at all
         (rule or ML) is retried through the ML classifier at a lower,
         highlight-only confidence threshold, since we already know a
         human flagged this column as relevant.

    Returns (col_map, col_source, col_confidence, highlighted_field_ids,
             unmatched_highlighted_headers) - the last is a list of
    {colIdx, header} for highlighted columns that STILL couldn't be
    mapped to any field even after the relaxed retry, for surfacing back
    to the user rather than silently dropping them.
    """
    col_map = dict(col_map)
    col_source = dict(col_source)
    col_confidence = dict(col_confidence)
    highlighted_field_ids = set()

    if not highlighted_cols:
        return col_map, col_source, col_confidence, highlighted_field_ids, []

    # 1. Conflict resolution
    by_field = {}
    for col_idx, fid in col_map.items():
        by_field.setdefault(fid, []).append(col_idx)

    for fid, cols in by_field.items():
        if fid in MULTI_COLUMN_AVERAGE_FIELDS:
            # These fields (e.g. O2out's Left/Right APH columns) are meant
            # to be averaged together deliberately (see
            # MULTI_COLUMN_AVERAGE_FIELDS / _dedupe_columns_per_field) —
            # don't drop one side just because the engineer only
            # highlighted the other. Still count the field as
            # "highlighted" if either column was, for the
            # highlighted-field bookkeeping below.
            if any(c in highlighted_cols for c in cols):
                highlighted_field_ids.add(fid)
            continue
        hi_cols = [c for c in cols if c in highlighted_cols]
        if hi_cols:
            highlighted_field_ids.add(fid)
        if hi_cols and len(hi_cols) < len(cols):
            for c in cols:
                if c not in hi_cols:
                    col_map.pop(c, None)
                    col_source.pop(c, None)
                    col_confidence.pop(c, None)

    # 2. Recovery for still-unmatched highlighted columns
    unmatched_highlighted = []
    retry_idx, retry_text = [], []
    for col_idx in sorted(highlighted_cols):
        if col_idx in col_map:
            continue
        hdr = headers[col_idx] if col_idx < len(headers) else None
        if hdr is None or not str(hdr).strip():
            continue
        if is_non_field_header(hdr):
            continue
        retry_idx.append(col_idx)
        retry_text.append(str(hdr))

    if use_ml and retry_text:
        clf = get_classifier()
        preds = clf.predict_batch(retry_text, threshold=HIGHLIGHTED_ML_THRESHOLD)
        for col_idx, hdr, pred in zip(retry_idx, retry_text, preds):
            fid, score, _ = pred
            if fid and fid not in NEVER_AUTO_DETECT and not _unit_conflicts_with_field(hdr, fid):
                col_map[col_idx] = fid
                col_source[col_idx] = 'ml_highlighted'
                col_confidence[col_idx] = round(score, 3)
                highlighted_field_ids.add(fid)
            else:
                unmatched_highlighted.append({'colIdx': col_idx, 'header': hdr})
    else:
        for col_idx, hdr in zip(retry_idx, retry_text):
            unmatched_highlighted.append({'colIdx': col_idx, 'header': hdr})

    return col_map, col_source, col_confidence, highlighted_field_ids, unmatched_highlighted


def _to_num(val):
    """Safely convert a cell value to float, or return None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip().replace(',', ''))
    except (ValueError, TypeError):
        return None


def _sym_to_field(sym):
    """Map a symbol string to a CENPEEP field id."""
    s = str(sym).strip()
    return SYM_MAP.get(s) or SYM_MAP_LOWER.get(s.lower())


def _match_tag_patterns(norm):
    """
    Secondary rule-based matcher for common plant/DCS tag phrasings that
    don't exactly match a LABEL_ALIASES entry verbatim but follow a
    recognizable token pattern (APH side-A/B tags, L_SIDE/R_SIDE, IN/OUT
    qualifiers, etc). Runs after the exact-alias lookup and before the ML
    fallback, so these well-known tag shapes resolve deterministically
    instead of depending on TF-IDF similarity (or a retrain) to catch them.

    `norm` is already lowercased and stripped of punctuation (see
    _label_to_field) — this only tokenizes on whitespace and checks for
    known word combinations.
    """
    tokens = set(norm.split())
    if not tokens:
        return None

    # "APH-A" / "APH-B" (side-A/side-B duct qualifiers) collapse into a
    # single "apha"/"aphb" token once punctuation is stripped during
    # normalization -- the hyphen disappears entirely rather than becoming
    # a space -- which silently defeated every 'aph' in tokens check below
    # for side-qualified real-plant headers like "FLUE GAS TEMP APH-A O/L,
    # DEG C" or "O2 APH-B O/L". Those never matched the APH in/out
    # patterns here and fell straight through to the ML fallback instead,
    # which then had no side-agnostic training example to latch onto.
    # Normalizing them back to a plain "aph" token makes side-qualified
    # and side-agnostic tag wording behave identically below.
    #
    # Real plant sheets also number the individual sensor on each side
    # (e.g. "APH-A1 O/L O2" and "APH-A2 O/L O2" for two separate O2 probes
    # on the A-side duct, or "O2 AT APH-A I/L-1"/"...I/L-2" for two inlet
    # probes) -- the sensor digit glues directly onto "apha"/"aphb" with no
    # separating punctuation left after normalization, so a plain equality
    # check against "apha"/"aphb" misses "apha1"/"apha2"/"aphb1"/"aphb2"
    # entirely, leaving those columns to fall through to the ML fallback
    # (which then had no reason to prefer O2in over O2out for an
    # inlet-side sensor numbered "-2", and mis-scored it onto O2out
    # instead). Match any side letter with an optional trailing sensor
    # number, not just the bare side letter, so every numbered probe on a
    # side folds down to the same side-agnostic "aph" token.
    tokens = {('aph' if re.fullmatch(r'apha\d*|aphb\d*', t) else t) for t in tokens}

    # Same hyphen-glued-side-qualifier problem as "APH-A"/"APH-B" above,
    # but for "ECO-L"/"ECO-R" (Left/Right duct O2 sensor "after the
    # economiser" -- real CSTPS header: "O2 IN FG AFT ECO-L, %" / "O2 IN FG
    # AFT ECO-R, %"). The hyphen disappears during normalization instead of
    # becoming a space, so "ECO-L" collapses onto the single token "ecol"
    # (not "eco" + "l"), which defeated the 'eco' in tokens check below.
    # Folded back to a plain "eco" token so side-qualified and
    # side-agnostic "after ECO" wording behave identically.
    tokens = {('eco' if t in ('ecol', 'ecor') else t) for t in tokens}

    # "AH" (Air Heater) is the same equipment CENPEEP/most tag wording
    # calls "APH" -- some sheets abbreviate it the short way instead (e.g.
    # "N2 At AH Outlet", "O2 At AH Outlet"). Folded to the same 'aph'
    # token used everywhere below so both spellings hit the same rules,
    # same precedent as the apha/aphb and ecol/ecor folds above.
    tokens = {('aph' if t == 'ah' else t) for t in tokens}

    # O2 in the flue gas "after"/"aft" the economiser (ECO) -- on this
    # plant's gas path this is physically the same point as the APH inlet
    # (see the "O2 AT ECO OUTLET" -> O2in training examples), just phrased
    # differently on some real sheets ("O2 IN FG AFT ECO-L, %"). Handled as
    # a deterministic rule rather than left to the ML fallback: char
    # n-grams in "...AFT ECO..." overlap heavily with the unrelated "FG
    # TEMPAFTERECO..." (Tgi) training examples and were winning that
    # match instead. Excludes "inlet"/"before" wording, which is the
    # different, further-upstream "O2 AT ECO INLET" reading (rejected
    # elsewhere as out-of-scope) rather than this after-ECO one.
    if 'o2' in tokens and 'eco' in tokens and ({'aft', 'after'} & tokens) \
            and not ({'inlet', 'inl', 'before'} & tokens):
        return 'O2in'

    # Unburnt-carbon-in-ash fields: "bottom ash"/"fly ash" (%), optionally
    # qualified with "unburnt(s)" — deliberately excludes phrasings that
    # also contain "total", since those mean the % that ash type makes up
    # of the TOTAL ash (Pba/Pfa) — a different field with near-identical
    # wording. See the LABEL_ALIASES note above for why this can't just be
    # a plain exact-string alias.
    if 'total' not in tokens:
        if {'bottom', 'ash'} <= tokens:
            return 'Cba'
        if {'fly', 'ash'} <= tokens:
            return 'Cfa'
    elif 'ash' in tokens:
        if 'bottom' in tokens:
            return 'Pba'
        if 'fly' in tokens:
            return 'Pfa'

    # O2 / CO at the APH inlet or outlet, in the "<reading> ... APH ...
    # <direction>" tag shape real DCS exports use (e.g. "O2 APH O/L",
    # "APH A OUTL GAS O2 CT", "APH B INL GAS O2 CT").
    if 'aph' in tokens:
        # Same glued-sensor-number problem as "apha"/"aphb" above, but for
        # the in/out qualifier itself -- a per-side sensor number can land
        # on "IL"/"OL" instead of (or in addition to) the side letter, e.g.
        # "O2 AT APH-A I/L-1"/"...I/L-2" normalizes each inlet reading's
        # qualifier to "il1"/"il2", not the bare "il" this membership
        # check was looking for. Match with an optional trailing digit so
        # every numbered probe's in/out qualifier is recognized the same
        # way, regardless of which sensor number it carries.
        is_out = any(re.fullmatch(r'(ol|out|outl|outlet|aft|after)\d*', t) for t in tokens)
        is_in = any(re.fullmatch(r'(il|in|inl|inlet|before|bef)\d*', t) for t in tokens)
        if 'o2' in tokens and is_out and not is_in:
            return 'O2out'
        if 'o2' in tokens and is_in and not is_out:
            return 'O2in'
        if 'co' in tokens and is_out and not is_in:
            return 'COout'
        if 'co' in tokens and is_in and not is_out:
            return 'COin'

        # Flue Gas / Primary Air / Secondary Air TEMPERATURE at the APH,
        # same tag shape (e.g. "FLUE GAS TEMP APH-A O/L, DEG C"). Only
        # fires when the header names which gas/air it is -- if none of
        # flue-gas/PA/SA is identifiable, this is ambiguous and falls
        # through to ML rather than guessing.
        if {'temp', 't'} & tokens:
            is_flue = bool({'flue', 'gas', 'fg'} & tokens)
            is_pa = bool({'pa', 'primary'} & tokens)
            is_sa = bool({'sa', 'secondary'} & tokens)
            if is_out and not is_in:
                if is_flue:
                    return 'Tgo'
                if is_pa:
                    return 'Tpao'
                if is_sa:
                    return 'Tsao'
            if is_in and not is_out:
                if is_flue:
                    return 'Tgi'
                if is_pa:
                    return 'Tpai'
                if is_sa:
                    return 'Tsai'

    # Secondary air arriving at the furnace after the APH — real plant tags
    # sometimes name this from the furnace's point of view ("FURNACE L_SIDE
    # INL SA T" / "FURNACE R_SIDE INL SA T") rather than the APH's point of
    # view, but it's the same physical reading CENPEEP calls "SA Temp Out"
    # (Tsao) — the secondary air has already passed through the APH by the
    # time it reaches the furnace inlet.
    if {'furnace', 'inl', 'sa'} <= tokens and ({'t', 'temp'} & tokens):
        return 'Tsao'

    # Main Steam (MS) Flow -- the primary CENPEEP steam-flow reading, and
    # the exact real-plant header wording on sheets like CSTPS's ("MS
    # FLOW, TPH"). This must be a deterministic RULE match, not left to
    # the ML fallback: rule matches always beat ML matches in
    # _dedupe_columns_per_field regardless of confidence score, but "MS
    # FLOW, TPH" only scored ~0.65 confidence against the Ffw training
    # examples (the ", TPH" unit suffix dilutes the char-ngram match),
    # while an unrelated "FEED FLOW" column on the same sheet scored
    # ~0.75 -- so FEED FLOW was winning the same-field tie-break and
    # silently overriding the real Main Steam Flow reading. Also covers
    # "M S FLOW" (space-separated) for the same reason, even though that
    # exact phrasing already has a dedicated training example -- pinning
    # it down as a rule too makes the outcome deterministic instead of
    # depending on a TF-IDF score staying above whatever else it's
    # competing against.
    if {'ms', 'flow'} <= tokens or {'m', 's', 'flow'} <= tokens:
        return 'Ffw'

    # Superheater-outlet Main Steam Pressure -- the physical point
    # CENPEEP's "Steam Pressure" (SP) reading wants (e.g. "SH O/L MS
    # PRESSURE-L"/"-R", "SH OUTLET MS PRESSURE"). Must be a deterministic
    # RULE match, not left to the ML fallback: a Hot/Cold Reheat pressure
    # column ("HRH STEAM PRESSURE-L") shares almost all of its vocabulary
    # ("STEAM"/"PRESSURE") with the handful of SP training phrases and was
    # winning the same-field tie-break at ~0.88 confidence, even though
    # HRH/CRH is a different, downstream point after the reheater rather
    # than the superheater-outlet reading SP actually means. (See also the
    # 'SP' entry in _unit_conflicts_with_field, which blocks HRH/CRH
    # columns from ever landing on SP even via ML.)
    # "PRESSURE-L"/"PRESSURE-R" glue their side letter directly onto
    # "pressure" once punctuation is stripped ("pressurel"/"pressurer"),
    # same glued-suffix problem as "apha"/"aphb" above -- so this checks
    # for a "pressure..." token by prefix rather than requiring an exact
    # "pressure" token.
    if {'sh', 'ms'} <= tokens and any(t.startswith('pressure') for t in tokens) \
            and ({'ol', 'out', 'outl', 'outlet'} & tokens):
        return 'SP'

    # Plant's own precomputed "Total PA/SA Flow" column (e.g. "Total PA
    # FLOW", "Total SA FLOW "). Summing the individual per-mill/per-duct
    # columns instead is a fragile fallback -- a mill legitimately cycling
    # off for part of the period leaves that column blank for those rows,
    # undercounting the sum (see MULTI_COLUMN_SUM_FIELDS) -- so when the
    # sheet already provides its own precomputed total, prefer it. This
    # must be a deterministic RULE match, not left to the ML fallback:
    # rule matches always beat ML matches in _dedupe_columns_per_field
    # regardless of confidence, and the "total" wording also keeps this
    # column's _base_tag_key too dissimilar from the per-mill columns'
    # (see _groups_are_siblings) for them to be blended in alongside it.
    if 'total' in tokens and 'flow' in tokens:
        if {'pa', 'primary'} & tokens:
            return 'Fpa'
        if {'sa', 'secondary'} & tokens:
            return 'Fsa'

    return None


# Field ids that are always a TEMPERATURE reading -- never legitimately a
# pressure/draft reading, no matter how a header's wording happens to
# score against the ML training set. Tfg/Tamb (BEE-2 Indirect's flue gas /
# ambient temperature) get the same guard as the CENPEEP APH duct
# temperatures for the same reason -- both share "FG"/"TEMP" wording with
# real draft-pressure tags.
TEMPERATURE_ONLY_FIELDS = {'Tgi', 'Tgo', 'Tpai', 'Tpao', 'Tsai', 'Tsao', 'Tfg', 'Tamb'}

# Unit tokens that mark a header as a PRESSURE/DRAFT reading (mmWC =
# millimeters water column, KSC/KG per CM2 = kg per sq cm) -- real plant
# sheets label draft-pressure tags this way right next to genuine
# temperature tags with very similar wording (e.g. "FG AH 9A OUT, MMWC"
# sitting next to "FLUE GAS TEMP APH-A O/L, DEG C"), and their shared
# words ("FG", "AH", "OUT") can fool the ML fallback into scoring the
# pressure column as a temperature field. This is a hard physical guard,
# not a fuzzy one: any header carrying one of these unit tokens can never
# resolve to a temperature-only field id, regardless of ML confidence.
# Tokens that mark a header as an O2/CO/CO2 GAS-ANALYSIS reading rather
# than a temperature -- real plant sheets place these right next to
# genuine temperature tags at the very same physical location ("FG O2 A
# AFTER APH" sitting next to "FLUE GAS TEMP APH-A O/L"), sharing enough
# words ("FG"/"AFTER"/"APH") that the ML fallback can score the O2%
# column as a temperature field instead — which is exactly what happened
# with an "FG O2 A AFTER APH" column landing on Tgo (flue gas temp after
# APH) at 0.878 confidence, averaging a handful of ~7 (percent!) readings
# in as if they were degrees C. Hard physical guard, same shape as
# PRESSURE_UNIT_TOKENS below: any header carrying one of these tokens can
# never resolve to a temperature-only field id, regardless of confidence.
GAS_ANALYSIS_TOKENS = {'o2', 'co2', 'co', 'oxygen', 'nox', 'sox', 'so2'}

# Words that mark a header as a TEMPERATURE reading -- the mirror image of
# GAS_ANALYSIS_TOKENS, used to stop the same kind of shared-wording mixup
# in the other direction (a genuine temperature column being ML-matched
# onto O2in/O2out/O2fg/COin/COout/COfg because it shares "after APH"-style
# process-location wording with a real gas-analysis training example).
# Deliberately just "temp"/"temperature" (not the bare single-letter "t",
# which is far too common a token/abbreviation elsewhere to safely use as
# a hard exclusion signal).
TEMPERATURE_WORD_TOKENS = {'temp', 'temperature'}

# Every field id physically measured as an O2/CO/CO2 percentage or ppm
# reading -- gets the reverse-direction guard described above.
GAS_ANALYSIS_ONLY_FIELDS = {'O2in', 'O2out', 'O2fg', 'COin', 'COout', 'COfg',
                             'CO2in', 'CO2out', 'CO2fg'}

# Unit tokens that mark a header as a PRESSURE/DRAFT reading (mmWC =
# millimeters water column, KSC/KG per CM2 = kg per sq cm) -- real plant
# sheets label draft-pressure tags this way right next to genuine
# temperature tags with very similar wording (e.g. "FG AH 9A OUT, MMWC"
# sitting next to "FLUE GAS TEMP APH-A O/L, DEG C"), and their shared
# words ("FG", "AH", "OUT") can fool the ML fallback into scoring the
# pressure column as a temperature field. This is a hard physical guard,
# not a fuzzy one: any header carrying one of these unit tokens can never
# resolve to a temperature-only field id, regardless of ML confidence.
PRESSURE_UNIT_TOKENS = {'mmwc', 'mmwcl', 'mmwg', 'ksc', 'kgcm2'}


def _unit_conflicts_with_field(header, fid):
    """True if header's own unit/quantity wording physically rules out fid."""
    norm = re.sub(r'[^a-z0-9 ]', '', str(header).lower())
    tokens = set(norm.split())
    if fid in TEMPERATURE_ONLY_FIELDS:
        return bool(tokens & PRESSURE_UNIT_TOKENS) or bool(tokens & GAS_ANALYSIS_TOKENS)
    if fid in GAS_ANALYSIS_ONLY_FIELDS:
        return bool(tokens & TEMPERATURE_WORD_TOKENS)
    if fid == 'SP':
        # Hot/Cold Reheat pressure ("HRH STEAM PRESSURE-L", "CRH STEAM
        # PRESS-R") is a different, downstream point after the reheater --
        # not the superheater-outlet Main Steam Pressure reading SP means
        # -- but shares almost all its vocabulary ("STEAM"/"PRESSURE")
        # with the SP training phrases, which let it out-score the real
        # SH-outlet reading via the ML fallback. Hard physical guard: any
        # header naming HRH/CRH can never resolve to SP, regardless of
        # confidence.
        return bool(tokens & {'hrh', 'crh'})
    return False


def _label_to_field(label):
    """Map a header label string to a CENPEEP field id."""
    norm = re.sub(r'[^a-z0-9 ]', '', str(label).lower().strip())
    # Direct symbol match first
    fid = _sym_to_field(label.strip())
    if fid:
        return fid
    fid = LABEL_ALIASES.get(norm)
    if fid:
        return fid
    return _match_tag_patterns(norm)


# ─── Strategy 1: Standard CENPEEP column layout ───────────────────────────────
def _parse_cenpeep_layout(rows):
    """
    Expects rows like:
      col0=Particulars, col1=UOM, col2=Symbol, col3=Formula/INPUT, col4=Value
    Returns (extracted_dict, raw_rows_list).
    """
    extracted = {}
    raw_rows = []
    design_md_seen = False
    design_ad_seen = False

    for row in rows:
        if len(row) < 5:
            continue
        particulars = row[0]
        uom = row[1]
        symbol = row[2]
        formula = row[3]
        value = row[4]

        if not symbol or value is None:
            continue

        is_input = isinstance(formula, str) and formula.strip().lower() == 'input'
        is_plain = (formula is None) and isinstance(value, (int, float))

        if not is_input and not is_plain:
            continue

        sym = str(symbol).strip()
        num = _to_num(value)
        if num is None:
            continue

        # Handle duplicate Md / Ad design symbols
        field_id = _sym_to_field(sym)
        if sym == 'Md':
            field_id = 'Md2' if design_md_seen else 'Md'
            design_md_seen = True
        if sym == 'Ad':
            field_id = 'Ad2' if design_ad_seen else 'Ad'
            design_ad_seen = True

        if not field_id or field_id in NEVER_AUTO_DETECT:
            continue

        extracted[field_id] = num
        raw_rows.append({
            'particulars': str(particulars) if particulars else sym,
            'uom': str(uom) if uom else '',
            'symbol': sym,
            'value': num,
        })

    return extracted, raw_rows


# ─── Header row detection (real sheets often bury the header a few rows down) ─
# Common short/unit-only tokens that carry no field-identifying meaning on
# their own. Real plant workbooks often stack a merged group-label cell
# ("O2% BEFORE AH AVG.") above a row of per-column sub-headers that just
# repeat the unit or side ("%", "Deg C", "LHS"/"RHS") -- read alone, the
# sub-header can't be matched (or gets matched wrong) to any field; see
# _is_bare_header_cell / _stitch_stacked_headers.
# Deliberately duplicates the letter/number/side tokens also captured by
# _SIDE_INDICATOR_TOKENS (defined later in this module, alongside the
# sibling-duct grouping logic it's used for) rather than referencing it
# directly, so this set is available immediately here without depending on
# module-level definition order.
_BARE_UNIT_TOKENS = {
    'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h',
    'l', 'r', 'i', 'ii', 'iii', '1', '2', '3', '4', '5', '6', '7', '8',
    'lhs', 'rhs', 'left', 'right', 'side',
    'pct', 'percent', 'deg', 'degc', 'degree', 'degrees', 'kg', 'mw', 'ppm',
    'avg', 'average', 'tph', 'hr', 'hrs', 'kcal', 'nm3', 'kj', 'no',
}


def _is_bare_header_cell(text):
    """
    True if `text` (a header cell) carries no field-identifying meaning by
    itself — either blank, or made up entirely of unit/side tokens (see
    _BARE_UNIT_TOKENS) with no other descriptive word. Such a cell is a
    sub-header that only makes sense combined with a group-label cell
    above it (see _stitch_stacked_headers).

    A single-letter/short cell that's actually a real CENPEEP SYMBOL on
    its own (e.g. a column literally headed "L" for Load, or "A" for Ash,
    per SYM_MAP) is deliberately NOT bare — the standard CenPeep column
    layout uses exactly these bare-looking single-letter symbols as real,
    complete headers, and they must keep resolving via the normal
    rule-based symbol lookup rather than being swept into the
    stacked-header rescue path (which only fires when a group label is
    found above it).
    """
    if text is None:
        return True
    if _sym_to_field(str(text).strip()):
        return False
    norm = re.sub(r'[^a-z0-9]+', ' ', str(text).lower()).strip()
    if not norm:
        return True
    tokens = set(norm.split())
    if tokens <= _BARE_UNIT_TOKENS:
        return True
    # A header cell that's purely numeric (e.g. "2.84") is never a
    # legitimate field label -- it's a value that landed on the header
    # row, most likely because the actual header text sits in a merged
    # cell/row above it (or the header-row heuristic picked a slightly
    # wrong row on an unusual sheet layout). Treating it as bare keeps it
    # off the ML fallback entirely rather than letting a stray number get
    # fuzzy-matched to whichever field its digits happen to resemble.
    if all(t.replace('.', '', 1).isdigit() for t in tokens):
        return True
    return False


def _forward_fill_merged_row(row):
    """
    Excel stores a merged cell's text only in its top-left cell — every
    other cell the merge visually covers reads back as blank. A group-label
    row like "O2% BEFORE AH AVG." spanning two columns (LHS/RHS) therefore
    often has that text in only the LEFTMOST of the two columns, with the
    other column's cell genuinely empty in the underlying data. Forward-
    filling (carrying the last non-blank cell rightward until the next
    non-blank cell) reconstructs what a person looking at the merged cells
    visually sees, so the RHS column can be stitched with the same group
    label as the LHS column instead of being left with nothing.
    """
    filled = []
    last = None
    for val in row:
        if val is not None and str(val).strip():
            last = val
        filled.append(last)
    return filled


def _stitch_stacked_headers(rows_above, header_row):
    """
    Combines a chosen header row with any group-label text sitting in the
    SAME column on the row(s) immediately above it, for columns whose own
    header cell is "bare" (see _is_bare_header_cell) — e.g. a merged
    "O2% BEFORE AH AVG." cell spanning several columns, with each column's
    own sub-header underneath just reading "%". Read alone, "%" can't be
    matched to any field; combined with the group label above it
    ("O2% BEFORE AH AVG. %"), it matches confidently and correctly.

    `rows_above` is the handful of rows immediately preceding the header
    row, in normal top-to-bottom order (closest row LAST) — searched
    nearest-first so the closest non-blank label wins, and forward-filled
    first (see _forward_fill_merged_row) so a merged group-label cell
    reaches every column it visually spans, not just its leftmost one.
    Only applied to columns whose header cell is bare — a column with its
    own real header text is left completely untouched, so this can never
    change how a sheet that already parses correctly today is read.
    """
    if not rows_above:
        return list(header_row)
    filled_rows_above = [_forward_fill_merged_row(r) for r in rows_above]
    stitched = []
    for col_idx, cell in enumerate(header_row):
        if not _is_bare_header_cell(cell):
            stitched.append(cell)
            continue
        label = None
        for above_row in reversed(filled_rows_above):
            if col_idx < len(above_row):
                val = above_row[col_idx]
                if val is not None and str(val).strip():
                    label = str(val).strip()
                    break
        if label:
            cell_text = str(cell).strip() if cell is not None else ''
            stitched.append(f"{label} {cell_text}".strip())
        else:
            stitched.append(cell)
    return stitched


# How many rows directly above the chosen header row to search for a
# group-label to stitch onto a bare sub-header cell (see
# _stitch_stacked_headers). Kept small and close to the header row on
# purpose — a title/banner row much further up the sheet is far more
# likely to be unrelated report text than a genuine merged group label.
STACKED_HEADER_LOOKUP_ROWS = 3


def _find_header_row(sample_rows, use_ml=True):
    """
    Scans the first few rows of a sheet and picks the one most likely to be
    a header row.

    Some real plant exports stack MULTIPLE header-shaped rows on top of each
    other before the data starts — e.g. row0 = human-readable description
    ("Main Steam Flow"), row1 = an aggregation label repeated across every
    column ("Hourly Average"), row2 = the raw PI/DCS tag code
    ("U6_MN_STM_TOT_FL"). All three are "mostly text, few numbers", so a
    pure text-density heuristic can't tell them apart — and can easily pick
    the least useful one (a tag-code row scores just as high on density as
    the readable row, sometimes higher if it has fewer blank/merged cells).

    So: first shortlist rows that look header-shaped (mostly text, few
    numbers), then, among the shortlist, actually try mapping each one to
    CENPEEP fields and pick whichever row yields the most matched fields.
    This directly optimizes for what the header row is FOR, instead of a
    proxy heuristic. Text-density score is only used as a tie-breaker.
    """
    candidates = []
    for i, row in enumerate(sample_rows):
        str_cells = sum(1 for c in row if isinstance(c, str) and c.strip())
        num_cells = sum(1 for c in row if isinstance(c, (int, float)))
        score = str_cells - num_cells
        if str_cells >= 3:
            candidates.append((i, score, row))

    if not candidates:
        return None

    best_idx, best_field_count, best_score = None, -1, -10 ** 9
    for i, score, row in candidates:
        col_map, _, _ = _map_columns_to_fields(row, use_ml=use_ml)
        field_count = len(set(col_map.values()))
        if field_count > best_field_count or (
            field_count == best_field_count and score > best_score
        ):
            best_idx, best_field_count, best_score = i, field_count, score

    # A field-count "winner" that only matched a couple of fields is weak
    # evidence -- genuine header rows on real CENPEEP/DCS exports typically
    # match many more fields than that (that's what makes the field-count
    # heuristic reliable enough to disambiguate stacked header rows in the
    # first place; see the docstring above). A row that only coincidentally
    # matches 1-2 fields is much more likely to be an ordinary DATA row
    # than a real header -- e.g. a totally unrelated cost-analysis table
    # whose "Formula" column happens to contain the bare letter "A" on one
    # row, which collides with the CENPEEP Ash symbol and got that row
    # picked as the "header", pulling a garbage average from an unrelated
    # column into the Ash field. Below this threshold, fall back to the
    # plain text-density heuristic (most text cells, fewest numeric cells)
    # among the same candidates instead of trusting the coincidence.
    MIN_HEADER_FIELD_COUNT = 3
    if best_field_count < MIN_HEADER_FIELD_COUNT:
        best_idx = max(candidates, key=lambda c: c[1])[0]

    return best_idx


# ─── Column → field mapping (rule-based alias lookup + ML fallback) ──────────
def _map_columns_to_fields(headers, use_ml=True, ml_threshold=DEFAULT_CONFIDENCE_THRESHOLD,
                            raw_headers=None):
    """
    Given a list of header strings (one per column), returns:
      col_map: {col_idx: field_id}
      col_source: {col_idx: 'rule' | 'ml'}
      col_confidence: {col_idx: float}   (1.0 for rule matches)
    Rule-based alias lookup runs first (cheap, exact); any column it can't
    place is handed to the ML classifier as a batch (fast — single vectorized
    call rather than one call per column).

    `headers` is normally the stacked-header-stitched list (see
    _stitch_stacked_headers) — a bare sub-header like "%" combined with the
    group label above it ("O2% BEFORE AH AVG."). If `raw_headers` (the
    UN-stitched originals) is also given, a stitched header is only ever
    used for the safe, literal rule/alias/symbol lookup (_label_to_field_exact)
    -- never for the fuzzy tag-pattern fallback or the ML classifier. Those
    two are the parts that judge a header by loose word/phrase similarity,
    and a stitched label built from a merged group-header can accidentally
    read as "close enough" to a totally different field's training phrases
    (e.g. "O2% BEFORE AH AVG. LHS" superficially resembling "flue gas temp
    before aph" and getting ML-matched to Tgi, a temperature field, instead
    of an O2 field). The raw, un-stitched header is what's actually run
    through the fuzzy/ML paths, so a bare column that isn't rescued by the
    literal stitched-alias lookup just stays unmatched rather than risking
    a wrong-field match — consistent with the rest of this module always
    preferring "leave it for the person to fill in" over guessing wrong.
    """
    if raw_headers is None:
        raw_headers = headers

    col_map = {}
    col_source = {}
    col_confidence = {}
    unmatched_idx = []
    unmatched_text = []

    for col_idx, hdr in enumerate(raw_headers):
        stitched_hdr = headers[col_idx] if col_idx < len(headers) else hdr
        effective_hdr = hdr if (hdr is not None and str(hdr).strip()) else None

        if effective_hdr is None or is_non_field_header(effective_hdr) or _is_bare_header_cell(effective_hdr):
            # Nothing usable in the raw cell itself. If stacked-header
            # stitching found a group label above it, give ONLY the safe,
            # literal exact/alias/symbol lookup a chance against the
            # combined text — see the docstring above for why this stays
            # off the fuzzy/ML paths.
            if stitched_hdr and str(stitched_hdr).strip() != (str(hdr).strip() if hdr else ''):
                fid = _label_to_field_exact(str(stitched_hdr))
                if fid and fid not in NEVER_AUTO_DETECT:
                    col_map[col_idx] = fid
                    col_source[col_idx] = 'rule'
                    col_confidence[col_idx] = 1.0
            continue

        fid = _label_to_field(str(effective_hdr))
        if fid and fid in NEVER_AUTO_DETECT:
            # Always-manual field (see NEVER_AUTO_DETECT) — treat the column
            # as unrecognized rather than auto-mapping it, and don't hand it
            # to the ML fallback either (a rule already matched it, just not
            # one we act on).
            continue
        if fid:
            col_map[col_idx] = fid
            col_source[col_idx] = 'rule'
            col_confidence[col_idx] = 1.0
        else:
            unmatched_idx.append(col_idx)
            unmatched_text.append(str(effective_hdr))

    if use_ml and unmatched_text:
        clf = get_classifier()
        preds = clf.predict_batch(unmatched_text, threshold=ml_threshold)
        for col_idx, (fid, score, matched_example) in zip(unmatched_idx, preds):
            if fid and fid not in NEVER_AUTO_DETECT and not _unit_conflicts_with_field(
                raw_headers[col_idx], fid
            ):
                col_map[col_idx] = fid
                col_source[col_idx] = 'ml'
                col_confidence[col_idx] = round(score, 3)

    col_map, col_source, col_confidence = _dedupe_columns_per_field(
        col_map, col_source, col_confidence, headers
    )

    return col_map, col_source, col_confidence


# Fields where two columns legitimately both belong to the same field id
# and should be AVERAGED together, rather than deduped down to one. This is
# the real-world "Left duct / Right duct" (or "A side / B side") dual-duct
# instrumentation pattern many plants use for APH readings — e.g. a sheet
# with BOTH "OXYGEN IN FLUE GAS(L)-AH O/L" and "OXYGEN IN FLUE GAS(R)-AH
# O/L" columns for Avg. Flue Gas O2 — APH Out (O2out): these are two
# separate physical sensors (left/right side of the air preheater), not a
# strict-vs-fuzzy duplicate match on the same reading, so both readings
# should be blended into the field's value instead of one being silently
# discarded by _dedupe_columns_per_field below.
#
# Every "Avg. Flue Gas ..." field (O2in/O2out/Tgi/Tgo) and every APH-duct
# air temperature field (Tpai/Tpao/Tsai/Tsao) is physically the same kind
# of dual-duct reading as O2out — a real plant sheet with, say,
# "APH-A I/L GAS TEMP" AND "APH-B I/L GAS TEMP" columns for Tgi should
# average both A/B-side sensors, not keep only whichever one happens to
# score marginally higher. Previously only O2out was listed here, so a
# sheet with two genuine Tgo columns (APH-A O/L FG TEMP + APH-B O/L FG
# TEMP) had one side silently dropped by the dedupe step below instead of
# being averaged in — the field label itself ("Avg. Flue Gas Temp — APH
# Out") already promises an average of both sides.
MULTI_COLUMN_AVERAGE_FIELDS = {
    'O2in', 'O2out', 'Tgi', 'Tgo', 'Tpai', 'Tpao', 'Tsai', 'Tsao',
    # SP ("Steam Pressure") is read at the superheater outlet, which (like
    # the APH/flue-gas ducts above) is commonly wired as separate L-side/
    # R-side sensors (e.g. "SH O/L MS PRESSURE-L" + "...-R") that should be
    # averaged together rather than one being silently dropped.
    'SP',
}

# Same "don't silently keep only one column" problem as
# MULTI_COLUMN_AVERAGE_FIELDS, but for fields where the physically correct
# combination across parallel ducts/mills is a SUM, not an average — e.g.
# Fpa ("Total Primary Air Flow") and Fsa ("Total Secondary Air Flow", see
# FIELD_LABELS) are the combined flow across every PA-to-mill duct / every
# SA duct, not the reading from whichever single duct happened to score
# best. A plant sheet with "PA FLOW TO MILL-A" .. "PA FLOW TO MILL-F" as
# six separate columns needs all six summed per row, then that per-row
# total averaged across rows — exactly mirroring how a human would total
# the mill flows for each timestamp before averaging over the period.
# These fields go through the same sibling-grouping machinery as
# MULTI_COLUMN_AVERAGE_FIELDS (see _dedupe_columns_per_field /
# _groups_are_siblings) but combine per-row with SUM instead of AVERAGE
# (see the row_sums / MULTI_COLUMN_SUM_FIELDS handling in
# _parse_raw_layout and _accumulate_row) and are not capped at two
# sibling groups, since a real plant can have more than two mills/ducts.
MULTI_COLUMN_SUM_FIELDS = {
    'Fpa', 'Fsa',
}

# Cap on how many sibling groups get kept for a MULTI_COLUMN_SUM_FIELDS
# field (e.g. one group per mill). Generous — real plants rarely exceed
# 8-10 mills/ducts for a single field — but still bounded so a pathological
# false-positive-heavy sheet can't pull in an unbounded number of columns.
MAX_SUM_FIELD_GROUPS = 12

# Words that mark WHICH physical duct/side/sensor a reading comes from
# (A/B/C.../LHS/RHS/left/right/1/2/3/...) rather than WHAT is being
# measured or WHERE in the process it sits. Two headers that differ only
# in these tokens are the same kind of reading on sibling ducts (e.g. "SA
# FLOW (LHS)" vs "SA FLOW (RHS)", or "PA FLOW TO MILL-A" vs "...MILL-B")
# and should be combined. Stripped out before comparing headers in
# _groups_are_siblings so a shared side-marker never counts as "these two
# headers are about the same thing" on its own, and a *different*
# equipment/location word (e.g. "mill" vs "aph", "before" vs "to") is what
# actually decides whether two columns are genuine siblings.
_SIDE_INDICATOR_TOKENS = {
    'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h',
    'l', 'r', 'i', 'ii', 'iii', '1', '2', '3', '4', '5', '6', '7', '8',
    'lhs', 'rhs', 'left', 'right', 'side',
}

# Minimum Jaccard overlap (after stripping side-indicator tokens) two
# candidate columns' headers must share to be treated as siblings of the
# same physical reading point — see _groups_are_siblings.
_SIBLING_TOKEN_OVERLAP_THRESHOLD = 0.5

# Words that mark a header as a plant-precomputed AGGREGATE of other
# columns (e.g. "Total PA FLOW") rather than one more sibling reading
# alongside them (e.g. "PA FLOW TO MILL-A"). See _groups_are_siblings.
_TOTAL_INDICATOR_TOKENS = {'total', 'overall', 'aggregate', 'sum'}


def _core_token_set(base_tag_key):
    """
    Splits an already-normalized _base_tag_key(...) string into its
    whitespace-separated tokens, with pure side-indicator tokens (see
    _SIDE_INDICATOR_TOKENS) removed — leaving just the tokens that
    describe WHAT/WHERE the reading is (e.g. "pa", "temp", "before",
    "aph", or "pa", "inlet", "temp", "to", "mill").
    """
    tokens = set(base_tag_key.split())
    return tokens - _SIDE_INDICATOR_TOKENS


def _groups_are_siblings(key_a, key_b):
    """
    True if two _base_tag_key(...) header keys describe the SAME physical
    reading point on two different ducts/sides/sensors (e.g. "flue gas
    temp before aph lhs" vs "...rhs"), rather than two genuinely different
    measurement points that merely both got mapped to the same field id
    (e.g. "pa temp before aph" vs "pa inlet temp to mill a" — sharing only
    "pa"/"temp" is NOT enough to call them the same reading).

    Measured as Jaccard overlap of each header's core tokens (side
    markers stripped) — high overlap means the headers differ only by
    which duct/side/sensor they're on; low overlap means they're
    describing different things that happen to share a field id, and
    should NOT be blended together.

    A header carrying a "total"/"overall" token is never a sibling of one
    that doesn't, regardless of overlap: a plant's own precomputed "Total
    PA/SA Flow" column is already an aggregate of the per-side/per-mill
    columns, so summing it alongside them (a MULTI_COLUMN_SUM_FIELDS
    field is combined per-row with SUM, not AVERAGE) would double-count
    it — the exact opposite of what "these describe the same reading, on
    different ducts" is meant to detect. High token overlap between e.g.
    "total sa flow" and "sa flow lhs" is precisely why this needs its own
    check rather than relying on the overlap score to sort it out.
    """
    tokens_a = _core_token_set(key_a)
    tokens_b = _core_token_set(key_b)
    if bool(tokens_a & _TOTAL_INDICATOR_TOKENS) != bool(tokens_b & _TOTAL_INDICATOR_TOKENS):
        return False
    if not tokens_a or not tokens_b:
        # No comparable tokens on one side (e.g. a header that was ALL
        # side-indicator words) — nothing to confirm a match against, so
        # don't treat it as a confident sibling.
        return tokens_a == tokens_b
    union = tokens_a | tokens_b
    if not union:
        return False
    overlap = len(tokens_a & tokens_b) / len(union)
    return overlap >= _SIBLING_TOKEN_OVERLAP_THRESHOLD


# Matches a single trailing sensor-number token at the very end of an
# already-whitespace-normalized, lowercased header (e.g. "...temp 2" ->
# strips " 2"). Deliberately anchored to the END ONLY, so a number that's
# part of the tag's middle (e.g. the "2" in "sec ar bx ilt 2 ar temp 1",
# short for "inlet-2 air") is left alone and doesn't get eaten — it's what
# keeps that tag distinct from an "ilt 1" / "inlet-1" tag on the other side.
_TRAILING_SENSOR_NUM_RE = re.compile(r'\s+\d+$')


def _is_opn_para_sheet(sheet_name):
    """
    True if `sheet_name` matches the real-plant "OPN PARA" (Operation
    Parameters) naming convention for the full-period master raw-data log
    -- e.g. "OPN PARA", "OPN_PARA", "OPN-PARA", "Operation Parameters".
    Used as a preference tiebreak when ranking raw tabular sheets; see the
    comment in the merge logic above.
    """
    norm = re.sub(r'[^a-z]', '', str(sheet_name).lower())
    return 'opnpara' in norm or 'operationparam' in norm or 'operatingparam' in norm


def _base_tag_key(header):
    """
    Reduces a header down to its "base tag" by stripping ONE trailing
    sensor-number suffix, so sibling sensors on the same physical duct
    (e.g. "APH-A I/L GAS TEMP-1" and "APH-A I/L GAS TEMP-2", or
    "APH-A O/L FG TEMP-1/-2/-3") normalize to the same key and are
    recognized as belonging together — while a genuinely different tag
    (e.g. the B-side duct, or an unrelated economiser/per-mill probe) keeps
    its own distinct key.
    """
    norm = re.sub(r'[^a-z0-9]+', ' ', str(header).lower()).strip()
    norm = re.sub(r'\s+', ' ', norm)
    # Some real plant sheets glue the per-sensor number directly onto the
    # side letter instead of trailing the whole header (e.g. "APH-A1 O/L
    # O2" / "APH-A2 O/L O2" for two separate O2 probes on the A-side duct,
    # vs the already-handled "...O2-1"/"...O2-2" trailing-suffix form).
    # Without folding "a1"/"a2"/"b1"/"b2" down to the bare side letter
    # first, each numbered probe gets its own single-column "group" below
    # and only ever the top TWO such groups survive the sibling-group cap
    # — silently dropping the other two probes instead of averaging all
    # four in. Folding first means "A1"+"A2" (and "B1"+"B2") key together
    # as one real duct-side group each, same as the trailing-suffix case.
    norm = ' '.join(
        re.sub(r'^([a-h]|l|r)\d+$', r'\1', tok) for tok in norm.split()
    )
    stripped = _TRAILING_SENSOR_NUM_RE.sub('', norm).strip()
    return stripped or norm


def _dedupe_columns_per_field(col_map, col_source, col_confidence, headers=None):
    """
    Keeps only the SINGLE best column for each field id, instead of letting
    every column that happens to map to the same field survive together —
    EXCEPT for fields listed in MULTI_COLUMN_AVERAGE_FIELDS (see above),
    where matched columns belonging to a genuine duct-side sensor group are
    deliberately kept so their readings get averaged together in
    _finalize_field_values.

    Real DCS/plant sheets routinely have one column with the exact/strict
    header (e.g. "MAIN STM FLOW COMP") AND one or more other columns whose
    wording is loosely/fuzzily similar (e.g. "T-FEED-FLOW", which just
    shares the words "feed"/"flow" with training phrasings like "Feed
    Water Flow"). Previously ALL such columns were kept, which corrupted
    results two ways: their values got averaged together in
    _finalize_field_values (blending a correct reading with an unrelated
    one), and whichever column happened to come LAST in column order won
    the "detected from" display — regardless of which one was actually the
    strict/confident match.

    Selection rule per field id, in order (fields in
    MULTI_COLUMN_AVERAGE_FIELDS skip this and instead run the grouped
    selection below):
      1. An exact rule match always beats an ML (fuzzy) match.
      2. Among same-source matches, higher confidence wins.
      3. Ties broken by earliest column index, for determinism.

    For MULTI_COLUMN_AVERAGE_FIELDS, columns are first grouped by
    _base_tag_key (which strips a trailing sensor-number suffix, e.g.
    "-1"/"-2"/"-3"), so that a duct side wired to 2 or 3 individual sensors
    (a very common real-plant setup — not just one reading per side) has
    ALL of its sensors kept and averaged in, not silently capped at one
    per side. The best TWO groups (by their best member's rank) are kept
    in full — this still protects against a wide sheet's other,
    loosely-related columns for the same field id (e.g. per-mill inlet
    temps for Tpai, or economiser-proxy temps for Tgi alongside a
    dedicated APH sensor), since those form their own separate group(s)
    and fall outside the top two. If headers aren't available, this falls
    back to the old best-two-columns behavior.

    Highlight-based conflict resolution (_apply_highlight_signal) still
    runs after this and can still override toward a human-highlighted
    column even if it isn't the strict winner here — that's a stronger,
    human-provided signal than header-text matching alone.
    """
    by_field = {}
    for col_idx, fid in col_map.items():
        by_field.setdefault(fid, []).append(col_idx)

    def rank(col_idx):
        # Lower is better: rule (0) beats ml (1); within a tier, higher
        # confidence is better (negated so sort ascending = best first);
        # earliest column index is the final tiebreaker.
        source_rank = 0 if col_source[col_idx] == 'rule' else 1
        return (source_rank, -col_confidence[col_idx], col_idx)

    new_map, new_source, new_confidence = {}, {}, {}
    for fid, cols in by_field.items():
        is_sum_field = fid in MULTI_COLUMN_SUM_FIELDS
        is_avg_field = fid in MULTI_COLUMN_AVERAGE_FIELDS
        if (is_sum_field or is_avg_field) and len(cols) > 1:
            group_cap = MAX_SUM_FIELD_GROUPS if is_sum_field else 2
            if headers is not None:
                groups = {}
                for col_idx in cols:
                    hdr = headers[col_idx] if col_idx < len(headers) else ''
                    key = _base_tag_key(hdr)
                    groups.setdefault(key, []).append(col_idx)

                def group_rank(group):
                    # Lower is better. Same source/confidence/index tiering
                    # as rank() above for the group's best member, but with
                    # group SIZE as the second criterion (before raw
                    # confidence): a 2- or 3-sensor group sharing a base tag
                    # (e.g. "...TEMP-1"/"-2"/"-3") is stronger evidence of a
                    # genuine duct-side sensor bank than a single column
                    # that happens to score a hair higher on its own — e.g.
                    # an exact-string training-data match at confidence 1.0
                    # for an unrelated single proxy column shouldn't be able
                    # to outrank and bump a real 2-sensor A-side/B-side pair
                    # out of the top two.
                    best = min(group, key=rank)
                    source_rank, neg_conf, idx = rank(best)
                    return (source_rank, -len(group), neg_conf, idx)

                ranked_groups = sorted(groups.values(), key=group_rank)
                ranked_keys = sorted(groups.keys(), key=lambda k: group_rank(groups[k]))

                # Only keep the best-ranked group unconditionally. Every
                # OTHER group must pass the sibling-token-overlap gate
                # against that top group (see _groups_are_siblings) before
                # it's allowed to join — this is what stops a column that
                # merely got independently classified into the same field
                # id (e.g. a per-mill "PA INLET TEMP TO MILL-A" reading
                # landing on Tpai alongside the real "PA TEMP BEFORE APH"
                # duct sensor) from being blended in just because it was
                # one of only two candidates. Genuine duct-side siblings
                # (LHS/RHS, A/B, ...) share all their non-side wording and
                # pass easily; unrelated columns that only coincidentally
                # share a field id don't.
                kept_keys = ranked_keys[:1]
                top_key = ranked_keys[0] if ranked_keys else None
                for key in ranked_keys[1:]:
                    if len(kept_keys) >= group_cap:
                        break
                    if top_key is not None and _groups_are_siblings(top_key, key):
                        kept_keys.append(key)
                keep_cols = [c for k in kept_keys for c in groups[k]]
            else:
                # No header text to group by — fall back to the old
                # best-two-individual-columns behavior (can't run the
                # sibling-token gate without header text to compare).
                keep_cols = sorted(cols, key=rank)[:group_cap]
            for col_idx in keep_cols:
                new_map[col_idx] = fid
                new_source[col_idx] = col_source[col_idx]
                new_confidence[col_idx] = col_confidence[col_idx]
            continue
        best = min(cols, key=rank)
        new_map[best] = fid
        new_source[best] = col_source[best]
        new_confidence[best] = col_confidence[best]

    return new_map, new_source, new_confidence


# ─── Row-count helpers (used to pick the right sheet when a field shows up
#     in more than one — e.g. an "Hourly" sheet AND a "Day Avg" sheet) ────────
_DATE_COL_HINTS = {'date', 'time', 'day', 'hour', 'hrs', 'hr', 'timestamp'}

# Excel's day-0, used to convert a raw numeric serial date (what xlrd hands
# back for date-formatted cells — it never auto-converts them the way
# openpyxl/calamine usually do) into a real calendar date. This is the
# standard 1900-date-system epoch that both Excel and Google Sheets use for
# the vast majority of real-world workbooks.
_EXCEL_EPOCH = datetime.datetime(1899, 12, 30)

# Hard cap on how many per-row dated snapshots we keep per sheet for the
# date-range feature (see _accumulate_row / _parse_raw_layout). A cap keeps
# memory/response size bounded even for a pathologically huge log sheet —
# every real CENPEEP hourly/15-min log seen in practice (a few thousand
# rows at most for a multi-month period) stays far under this.
MAX_DATED_ROWS_PER_SHEET = 20000


def _parse_date_cell(val):
    """
    Best-effort parse of a single cell value into a plain calendar date,
    regardless of how the source workbook represents it:
      - a native Python datetime/date (what openpyxl/calamine usually
        already hand back for a date-formatted cell)
      - a raw Excel serial number (what xlrd hands back — it never
        auto-converts date cells, see _iter_sheet_rows_streamed)
      - free text in more or less any common layout ("12-01-2024",
        "1/12/24", "2024-01-12", "1 Jan 2024", etc)
    Returns None if the value can't be read as a date at all (blank cell,
    or a bare time-of-day fraction with no date component) — callers treat
    that row as having an unknown date, not as an error.
    """
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, (int, float)):
        if val < 1:  # bare time-of-day serial (e.g. 0.5 = 12:00) — no date
            return None
        try:
            return (_EXCEL_EPOCH + datetime.timedelta(days=float(val))).date()
        except (OverflowError, ValueError):
            return None
    text = str(val).strip()
    if not text or not HAS_DATEUTIL:
        return None
    try:
        # dayfirst=True matches the dd-mm-yyyy convention most real plant/
        # DCS exports use; dateutil still resolves unambiguous formats
        # (yyyy-mm-dd, "Jan 2024") correctly regardless of this flag.
        return _dateutil_parser.parse(text, dayfirst=True, fuzzy=False).date()
    except (ValueError, OverflowError, TypeError):
        return None


def _find_date_col_idx(headers):
    """
    Locate a date/time-like column among the headers, purely to gauge how
    many real data rows a sheet has. This is NOT used for field extraction
    (date/time columns are never mapped to a CENPEEP field — see
    NON_FIELD_HEADERS) — it's only a row-counting yardstick so we can tell
    a genuine hourly log sheet (many populated date rows) apart from a
    daily/monthly summary sheet (few rows) when both sheets happen to
    produce a value for the same field.
    """
    for idx, h in enumerate(headers):
        if h is None:
            continue
        norm = re.sub(r'[^a-z0-9]+', ' ', str(h).lower()).strip()
        if set(norm.split()) & _DATE_COL_HINTS:
            return idx
    return None


def _find_date_col_idx_by_data(data_rows, col_map, sample_size=40):
    """
    Fallback for sheets where the row that actually got picked as the
    header (see _find_header_row — it's chosen by which row maps to the
    most CENPEEP fields, not by which row says "Date") doesn't itself
    carry a literal date-ish label. Real plant workbooks often stack a
    grouping row above the real header — e.g. a lab/quality sheet with
    "Date" on one row and the real sub-headers ("M%", "TM%", "A%", "FC%",
    "Fly Ash", "Bottom Ash", ...) one row below it — so _find_date_col_idx
    (header-text-only) finds nothing there even though the sheet
    genuinely has one row per day.
    Without a recognized date column, that sheet's rows can never be
    placed into any date range: every field sourced from it is stuck
    using its whole-sheet average forever, even after a person picks a
    specific date range on the calculator — which is exactly the
    "took the average of the whole [sheet], not just the selected
    dates" bug for fields like Fixed Carbon, GCV, Moisture, Volatile
    Matter, and Unburnt Carbon in Fly/Bottom Ash on this kind of sheet.
    So: scan a sample of the actual DATA rows (not headers) and look for
    a column where every sampled non-blank cell parses as a real date
    (_parse_date_cell) AND the parsed dates never run backwards row to
    row (a genuine chronological log is always sorted, or at least never
    goes in reverse; an ordinary measurement column of similar-looking
    numbers essentially never satisfies both). Only a column clearing
    both bars is treated as a date column — this deliberately stays
    conservative so a coincidentally date-like numeric column is never
    mistaken for one.
    Only rows that actually contain a real reading in at least one mapped
    field column are considered — this skips spacer / section-label rows
    (e.g. a literal "Without THERMACT" text label some plant sheets
    insert between blocks of dated rows) so a stray label can't
    disqualify an otherwise perfectly good date column just because it
    doesn't parse as a date.
    Returns the column index, or None if nothing qualifies (callers then
    behave exactly as if no date column exists — same as before).
    """
    if not data_rows:
        return None
    exclude_cols = set(col_map.keys())
    candidate_rows = [
        row for row in data_rows
        if any(_to_num(row[c]) is not None for c in col_map if c < len(row))
    ]
    sample = candidate_rows[:sample_size]
    ncols = max((len(r) for r in sample), default=0)
    best_idx, best_ratio = None, 0.0
    for c in range(ncols):
        if c in exclude_cols:
            continue
        parsed = []
        non_blank = 0
        for row in sample:
            if c >= len(row):
                continue
            val = row[c]
            if val is None or (isinstance(val, str) and not val.strip()):
                continue
            non_blank += 1
            d = _parse_date_cell(val)
            if d is not None:
                parsed.append(d)
        # Require EVERY non-blank sampled cell in the column to parse as a
        # date, and at least a handful of samples to trust the signal.
        if non_blank < 3 or len(parsed) != non_blank:
            continue
        if any(parsed[i] > parsed[i + 1] for i in range(len(parsed) - 1)):
            continue
        ratio = len(parsed) / non_blank
        if ratio > best_ratio:
            best_idx, best_ratio = c, ratio
    return best_idx


def _row_has_data(row, date_col_idx, col_map):
    """
    True if this row should count as a real reading, not a blank/spacer row.

    If we found a date/time column, a row only counts if that cell is
    actually populated — this is the direct implementation of "skip blank
    rows, don't count them, and don't treat them as a reading of 0".
    If no date column was found on this sheet, fall back to "does any
    mapped field column have a real number in this row".
    """
    if date_col_idx is not None:
        val = row[date_col_idx] if date_col_idx < len(row) else None
        return val is not None and str(val).strip() != ''
    for col_idx in col_map:
        val = row[col_idx] if col_idx < len(row) else None
        if _to_num(val) is not None:
            return True
    return False


def _count_populated_data_rows(data_rows, date_col_idx, col_map):
    """Counts real (non-blank) data rows in a sheet — see _row_has_data."""
    return sum(1 for row in data_rows if _row_has_data(row, date_col_idx, col_map))


# ─── Strategy 2: Raw tabular layout (header row + data rows), ML-augmented ───
def _parse_raw_layout(rows, use_ml=True, highlight_map=None):
    """
    Handles sheets where some early row is a header row with field names /
    symbols / free-text labels, and subsequent rows are data.

    Multiple data rows = multiple readings → averaged automatically.
    Unmatched headers are sent through the ML classifier as a fallback.
    `highlight_map`, if given, is {rowIdx: {colIdx, ...}} for this sheet's
    highlighted header cells (see _scan_header_highlights) — used to prefer
    highlighted columns on a field conflict and to recover highlighted
    columns the rule/ML pass missed (see _apply_highlight_signal).
    Returns (extracted_dict, raw_rows_list, sheet_summary, col_meta,
             data_row_count, unmatched_highlighted).
    """
    sample = rows[:HEADER_SCAN_ROWS]
    header_row_idx = _find_header_row(sample, use_ml=use_ml)

    if header_row_idx is None:
        return {}, [], {}, {}, 0, [], []

    raw_headers = rows[header_row_idx]
    headers = _stitch_stacked_headers(
        rows[max(0, header_row_idx - STACKED_HEADER_LOOKUP_ROWS):header_row_idx],
        raw_headers,
    )
    data_rows = rows[header_row_idx + 1:]

    col_map, col_source, col_confidence = _map_columns_to_fields(
        headers, use_ml=use_ml, raw_headers=raw_headers
    )

    highlighted_cols = (highlight_map or {}).get(header_row_idx, set())
    col_map, col_source, col_confidence, highlighted_field_ids, unmatched_highlighted = (
        _apply_highlight_signal(col_map, col_source, col_confidence, headers,
                                 highlighted_cols, use_ml=use_ml)
    )

    if not col_map:
        return {}, [], {}, {}, 0, unmatched_highlighted, []

    date_col_idx = _find_date_col_idx(headers)
    if date_col_idx is None:
        # Header text alone found nothing (e.g. "Date" sits on a stacked
        # grouping row above this one) — fall back to recognizing a date
        # column from the actual data. See _find_date_col_idx_by_data.
        date_col_idx = _find_date_col_idx_by_data(data_rows, col_map)
    data_row_count = _count_populated_data_rows(data_rows, date_col_idx, col_map)

    # Collect numeric values per field across all data rows. Blank / non-
    # numeric cells (_to_num returns None for these) are skipped outright —
    # they are never counted as 0, so they can't drag the average down.
    #
    # Alongside the flat per-field lists (used for the overall/undated
    # average, unchanged from before), also keep a per-ROW dated snapshot
    # of whatever fields that row actually populated — this is what lets
    # the frontend later re-average just the rows inside a chosen date
    # range (see MAX_DATED_ROWS_PER_SHEET) without re-uploading/re-parsing
    # the file. A row only gets a snapshot if it actually has at least one
    # numeric reading; its date comes from date_col_idx if one was found,
    # else None (row is still usable for the undated overall average, just
    # can't be placed into any date-range bucket).
    field_values = {fid: [] for fid in col_map.values()}
    dated_rows = []
    for row in data_rows:
        row_vals = {}
        row_sums = {}
        row_counts = {}
        for col_idx, fid in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            num = _to_num(val)
            if num is not None and not _reading_in_physical_range(fid, num):
                num = None
            if num is not None:
                if fid not in MULTI_COLUMN_SUM_FIELDS:
                    field_values[fid].append(num)
                # See the matching comment in _accumulate_row: a field can
                # have more than one column on the same row (e.g. O2in's
                # Left/Right duct columns) -- average them for this row's
                # dated snapshot instead of the second column overwriting
                # the first, which otherwise left the per-date-range
                # average (used once a person picks a date on the
                # calculator) silently built from only one of the columns.
                # MULTI_COLUMN_SUM_FIELDS (e.g. Fpa/Fsa — "Total" duct
                # flow, see the MULTI_COLUMN_SUM_FIELDS comment) instead
                # want that row's columns SUMMED, not averaged.
                row_sums[fid] = row_sums.get(fid, 0.0) + num
                row_counts[fid] = row_counts.get(fid, 0) + 1
        for fid, total in row_sums.items():
            if fid in MULTI_COLUMN_SUM_FIELDS:
                row_vals[fid] = total
                field_values[fid].append(total)
            else:
                row_vals[fid] = total / row_counts[fid]
        if row_vals and len(dated_rows) < MAX_DATED_ROWS_PER_SHEET:
            row_date = None
            if date_col_idx is not None and date_col_idx < len(row):
                row_date = _parse_date_cell(row[date_col_idx])
            dated_rows.append({
                'date': row_date.isoformat() if row_date else None,
                'values': row_vals,
            })

    extracted, raw_rows, sheet_summary = _finalize_field_values(field_values)
    col_meta = {
        col_idx: {
            'fieldId': fid,
            'header': str(headers[col_idx]),
            'source': col_source[col_idx],
            'confidence': col_confidence[col_idx],
            'highlighted': col_idx in highlighted_cols,
        }
        for col_idx, fid in col_map.items()
    }
    return (extracted, raw_rows, sheet_summary, col_meta, data_row_count,
            unmatched_highlighted, dated_rows)


# Hard physical bounds for fields whose real-world value can never fall
# outside a known range, no matter how a header got matched to them or
# what the sheet's own numbers say — a backstop against a bad match
# silently producing a nonsense average that then poisons every formula
# downstream (e.g. BEE-2 Indirect's EA = O2fg/(21-O2fg)*100 divides by a
# term that goes negative, and stays negative, the moment O2fg is fed
# anything at or above 21 — which is exactly what happened with a "SOX IN
# FLUE GAS" column getting ML-matched onto O2fg with an average of ~1794:
# EA/AAS/m all went negative, several losses went negative with them, and
# Boiler Efficiency came out above 100%). O2/CO2 are physically capped by
# air's own composition (~21% O2, ~20% CO2 is already an extreme flue-gas
# reading); CO is generous at up to 5% (50,000 ppm) since some sheets
# report it in ppm and some in %, and this only needs to catch clearly
# impossible values, not fine-tune plausible ones.
#
# The APH/economizer-duct temperature fields get the same backstop for a
# different real failure mode: a genuine DCS logging fault (a sensor
# transducer glitch, or a broken formula cell in the source workbook)
# occasionally logs one or two rows as an enormous garbage number (e.g.
# -176,928,736 sitting among thousands of otherwise sane ~260 deg C
# readings for the very same column) — a single such row is enough to
# drag a whole quarter's average to something like -43,902 deg C. -50..900
# deg C comfortably covers everything from a cold-start ambient reading to
# the hottest APH-outlet duct gas, so genuine plant readings are never
# affected — only DCS-fault-magnitude garbage gets rejected.
PHYSICAL_RANGE_FIELDS = {
    'O2fg': (0, 21), 'O2in': (0, 21), 'O2out': (0, 21),
    'CO2fg': (0, 21), 'CO2in': (0, 21), 'CO2out': (0, 21),
    'COfg': (0, 50000), 'COin': (0, 50000), 'COout': (0, 50000),
    'Tgi': (-50, 900), 'Tgo': (-50, 900), 'Tfg': (-50, 900),
    'Tpai': (-50, 900), 'Tpao': (-50, 900),
    'Tsai': (-50, 900), 'Tsao': (-50, 900),
}


def _reading_in_physical_range(fid, num):
    """
    True if a single raw reading `num` for field `fid` is within
    PHYSICAL_RANGE_FIELDS for it (or `fid` has no registered bound, in
    which case anything numeric is accepted). Used to drop individual
    DCS-fault-magnitude garbage readings AT ACCUMULATION TIME — before
    they ever reach an average — rather than only catching it after the
    fact once it's already dragged the whole field's average outside a
    sane range (see the PHYSICAL_RANGE_FIELDS comment above). Rejecting
    per-reading like this means the rest of a column's thousands of
    genuinely good readings still produce a correct average instead of
    the entire field being discarded over one or two bad rows.
    """
    bounds = PHYSICAL_RANGE_FIELDS.get(fid)
    if not bounds:
        return True
    return bounds[0] <= num <= bounds[1]


def _finalize_field_values(field_values):
    """Average collected numeric readings per field; build raw_rows + summary.
    Drops any field whose averaged value falls outside PHYSICAL_RANGE_FIELDS
    for it — see the comment above — rather than returning it as a
    confidently "detected" value that's actually impossible."""
    extracted = {}
    raw_rows = []
    sheet_summary = {}
    for fid, vals in field_values.items():
        if not vals:
            continue
        avg = statistics.mean(vals)
        bounds = PHYSICAL_RANGE_FIELDS.get(fid)
        if bounds and not (bounds[0] <= avg <= bounds[1]):
            continue
        extracted[fid] = avg
        sheet_summary[fid] = {'count': len(vals), 'values': vals[:50], 'average': avg}
        raw_rows.append({
            'particulars': fid, 'uom': '', 'symbol': fid,
            'value': avg, 'readings': len(vals),
        })
    return extracted, raw_rows, sheet_summary


# ─── Strategy 3: Generic labeled-row layout ───────────────────────────────────
# A looser variant of the strict CENPEEP column layout (Strategy 1) for
# sheets that are still fundamentally "one row per parameter, with a Symbol
# column" but don't match Strategy 1's fixed col0..col4 positions, and/or
# report several reading columns side-by-side (e.g. one per week/period)
# instead of a single Value column. Seen on real non-CENPEEP-authored
# efficiency sheets (e.g. plants following a different national methodology)
# that are otherwise structurally identical in spirit.
#
# IMPORTANT: this only ever runs when Strategies 1 and 2 BOTH find nothing
# (see _parse_sheet_rows / _parse_sheet_chunked below) — so it can never
# steal or change how a sheet that already parses correctly today is
# handled. Its symbol aliases are also kept in a table separate from
# SYM_MAP/LABEL_ALIASES for the same reason: they only ever apply inside
# this fallback, never inside Strategy 1 or Strategy 2's matching.
GENERIC_ROW_LAYOUT_SYMBOL_HEADERS = {'symbol', 'ky hieu'}
GENERIC_ROW_LAYOUT_UNIT_HEADERS = {'unit', 'uom', 'don vi'}
GENERIC_ROW_LAYOUT_PARTICULARS_HEADERS = {
    'particulars', 'parameter', 'description', 'thong so',
}
GENERIC_ROW_LAYOUT_SKIP_HEADERS = {
    'data collection method', 'method', 'stt', 'no', 'no.', 'sr no', 'sr no.',
    'phuong phap thu thap so lieu',
}
GENERIC_ROW_LAYOUT_HEADER_SCAN_ROWS = HEADER_SCAN_ROWS * 3

# Extra symbol aliases seen on non-CENPEEP-standard sheets that otherwise
# follow this same "row per parameter" shape (e.g. an EVN/Vietnamese
# boiler-efficiency methodology sheet using Ne/Qfw/tpa/tsa/... instead of
# CENPEEP's L/Ffw/Tpai/Tsai/...). Deliberately NOT merged into SYM_MAP.
GENERIC_ROW_LAYOUT_SYMBOL_ALIASES = {
    'ne': 'L',                  # Unit Load
    'qfw': 'Ffw', 'qsh': 'Ffw',  # Feed Water / Main Steam Flow Rate
    'tpa': 'Tpai',               # Primary Air Temp - AH inlet
    'tsa': 'Tsai',               # Secondary Air Temp - AH inlet
    'qpa': 'Fpa',                # Primary Air Flow Rate to AH
    'qsa': 'Fsa',                # Secondary Air Flow Rate to AH
    'tg14': 'Tgi',               # Flue Gas Temp - AH inlet
    'tg15': 'Tgo',               # Flue Gas Temp - AH outlet
    'o2ou': 'O2out', 'o2out': 'O2out',   # Flue Gas O2 - AH outlet
    'wc': 'M',                   # Moisture content of test coal
    'vc': 'VM',                  # Volatile matter of test coal
    'ac': 'A',                   # Ash content of test coal
    'qc': 'GCV',                 # Gross calorific value of test coal
    # Bare "N" for elemental coal Nitrogen (ultimate analysis notation --
    # C/H/N/O/S -- as opposed to CENPEEP's own "N2" symbol for the same
    # quantity). Kept here rather than in SYM_MAP because "N" alone is too
    # generic a symbol to trust outside this row-per-parameter layout.
    # Also reachable from an apostrophe-qualified variant ("N'", an
    # air-dried-basis intermediate some sheets carry alongside the
    # as-fired value) since the alnum-only normalization below strips the
    # apostrophe the same way -- _row_is_flue_gas_context /
    # the "(+)"-marked-row preference in _parse_generic_row_layout is what
    # picks the right one of the two when both are present.
    'n': 'N2',
}

# Symbols that ONLY ever appear as CALCULATED OUTPUTS of the boiler-
# efficiency computation itself (the individual loss-category breakdown),
# never as a raw plant/lab INPUT reading. A Strategy-3 ("Symbol" column)
# sheet whose Symbol column contains several of these is a derived/
# computed results sheet -- e.g. a "BOILER EFFICIENCY CALCULATION" /
# reference sheet the user built to cross-check the calculator's own
# output against -- even though it ALSO happens to re-list the input
# parameters that fed that calculation (O2in, O2out, Tgi, Tgo, GCV, Cba,
# Cfa, ... which legitimately match SYM_MAP). Genuine raw DCS/lab export
# files don't carry these loss-breakdown rows, so their presence is a
# reliable signal that this sheet is a calculated/reference sheet, not a
# raw input source -- see _parse_generic_row_layout / the merge logic in
# _extract_from_workbook, which rank such sheets LAST (after every other
# sheet, including big raw log sheets) instead of letting them win first
# just because they happen to have a literal "Symbol" column.
CALCULATED_OUTPUT_ONLY_SYMBOLS = {
    'ldg', 'luc', 'lmf', 'lhf', 'lco', 'lma', 'lrad',
}

# Field ids that are physically a property of the COAL/FUEL itself
# (proximate/ultimate analysis) -- never a legitimate flue-gas/APH-duct
# process reading. A Strategy-3 sheet's Symbol column is trusted blindly
# by field id, with zero regard for what the row's own Description says
# or which section of the sheet it's in -- which is exactly how a
# flue-gas reading whose symbol happens to collide with a CENPEEP
# coal-composition symbol (e.g. this workbook's "N2 At AH Outlet" duct
# reading vs CENPEEP's own "N2" symbol for coal Nitrogen) gets mistaken
# for the fuel-analysis value. See _row_is_flue_gas_context below.
GENERIC_ROW_LAYOUT_COMPOSITION_FIELDS = {
    'M', 'A', 'VM', 'FC', 'GCV', 'C', 'S', 'H2', 'N2', 'O2f',
}

# Description-text tokens that mark a row as a flue-gas/APH-duct process
# reading. A row is only treated as flue-gas context when it has BOTH a
# location word (APH/flue/stack/...) AND an in/out direction word --
# "Ash" alone or "Nitrogen content Of Coal" never trip this, only
# something shaped like "<reading> At AH Outlet" does.
_FLUE_GAS_LOCATION_TOKENS = {'aph', 'flue', 'stack', 'duct', 'economiser', 'economizer'}
_FLUE_GAS_DIRECTION_TOKENS = {'outlet', 'inlet', 'ol', 'il', 'out', 'in'}
# If the description itself names the fuel, it can never be a flue-gas
# reading no matter what other words it contains -- overrides the above.
_FUEL_CONTEXT_TOKENS = {'coal', 'fuel'}

# A trailing parenthetical qualifier on a Strategy-3 symbol (e.g. "A(f)",
# "GCV(f)") -- seen on non-CENPEEP-authored sheets that suffix a
# proximate-analysis symbol with its basis ("(f)" = as-fired) rather than
# using CENPEEP's own bare symbol for the as-fired value. Stripped before
# the SYM_MAP lookup so "A(f)" still resolves to the same field id "A"
# does. Naive punctuation-stripping alone can't do this: removing just
# the parentheses from "A(f)" leaves "Af", not "A" -- this removes the
# whole trailing "(...)" group instead.
_TRAILING_SYMBOL_QUALIFIER_RE = re.compile(r'\s*\([^()]*\)\s*$')


def _row_is_flue_gas_context(description):
    """
    True if `description` (a Strategy-3 row's Particulars/Description
    text) reads as a flue-gas/APH-duct process reading rather than a
    coal/fuel lab-analysis result -- see GENERIC_ROW_LAYOUT_COMPOSITION_FIELDS
    above. Used to reject a symbol match for a composition-only field id
    whose row is actually measuring something in the gas path.
    """
    norm = re.sub(r'[^a-z0-9]+', ' ', str(description).lower()).strip()
    tokens = set(norm.split())
    tokens = {('aph' if t == 'ah' else t) for t in tokens}
    if tokens & _FUEL_CONTEXT_TOKENS:
        return False
    return bool(tokens & _FLUE_GAS_LOCATION_TOKENS) and bool(tokens & _FLUE_GAS_DIRECTION_TOKENS)


def _generic_sym_to_field(sym):
    """Like _sym_to_field, plus the Strategy-3-only alias table above."""
    fid = _sym_to_field(sym)
    if fid:
        return fid
    stripped = _TRAILING_SYMBOL_QUALIFIER_RE.sub('', str(sym)).strip()
    if stripped and stripped != str(sym).strip():
        fid = _sym_to_field(stripped)
        if fid:
            return fid
    norm = re.sub(r'[^a-z0-9]', '', str(sym).lower().strip())
    return GENERIC_ROW_LAYOUT_SYMBOL_ALIASES.get(norm)


def _find_generic_header_row(sample_rows):
    """
    Finds a row containing a cell that reads exactly 'Symbol' (or a
    localized equivalent) - the one reliable, position-independent anchor
    for this layout. Returns (rowIdx, symbolColIdx) or (None, None).
    """
    for i, row in enumerate(sample_rows):
        for c_idx, cell in enumerate(row):
            if cell is None:
                continue
            norm = re.sub(r'[^a-z ]', '', str(cell).lower().strip())
            if norm in GENERIC_ROW_LAYOUT_SYMBOL_HEADERS:
                return i, c_idx
    return None, None


def _parse_generic_row_layout(rows, sheet_name=None):
    """
    Strategy 3 - see module comment above. Every row below the detected
    header is one parameter: its symbol lives in the Symbol column, and any
    other column on that row (except ones whose OWN header marks them as
    Unit/Method/serial-number/particulars text) is treated as a reading for
    that parameter. Multiple reading columns on the same row (e.g. one per
    week) are averaged together, same spirit as Strategy 2 averaging
    multiple DATA ROWS for one column.

    A row's Symbol cell is trusted first, but two extra checks now run
    before a row's numbers are accepted for a field:
      1. Flue-gas-context guard: if the field id is a coal/fuel
         COMPOSITION field (see GENERIC_ROW_LAYOUT_COMPOSITION_FIELDS) but
         the row's own Description reads as a flue-gas/APH-duct process
         reading (see _row_is_flue_gas_context), the row is rejected for
         that field -- a Symbol match alone isn't enough evidence when the
         row is plainly describing something else physically.
      2. "(+)"-marked-row preference: some sheets carry more than one row
         for the same field id (e.g. an as-fired value AND an air-dried-
         basis intermediate used only to derive it) and mark, via their own
         legend, which one is "actually used" -- typically with a "(+)" in
         the Description. When at least one surviving candidate row for a
         field has "+" in its description, only those marked row(s) are
         used; unmarked duplicates are ignored instead of being blended in.
         If none is marked, every surviving row is kept and averaged
         together, same as before -- this only changes behavior for sheets
         that actually use this marking convention.
    A row whose Symbol doesn't resolve to a field is also tried against its
    own Description text (same rule/tag matching Strategy 2 uses for
    headers), which recovers fields identifiable from wording alone even
    though the Symbol column doesn't carry a recognized abbreviation for
    them (e.g. a bare "O2"/"CO" symbol next to a "... At AH Outlet"
    description).

    Returns (extracted_dict, raw_rows_list, sheet_summary, col_meta,
             data_row_count, is_calculated_only).
    is_calculated_only is True when the Symbol column contains
    CALCULATED_OUTPUT_ONLY_SYMBOLS (see constant above) -- i.e. this looks
    like a derived boiler-efficiency results/reference sheet rather than a
    genuine raw-input sheet. Callers should rank such a sheet last.
    """
    sample = rows[:GENERIC_ROW_LAYOUT_HEADER_SCAN_ROWS]
    header_row_idx, symbol_col = _find_generic_header_row(sample)
    if header_row_idx is None:
        return {}, [], {}, {}, 0, False

    header_row = rows[header_row_idx]
    value_cols = []
    particulars_col = None
    for c_idx, cell in enumerate(header_row):
        if c_idx == symbol_col:
            continue
        text = '' if cell is None else str(cell).strip()
        norm = re.sub(r'[^a-z ]', '', text.lower())
        if text and norm in GENERIC_ROW_LAYOUT_PARTICULARS_HEADERS:
            if particulars_col is None:
                particulars_col = c_idx
            continue
        if text and (
            norm in GENERIC_ROW_LAYOUT_UNIT_HEADERS
            or norm in GENERIC_ROW_LAYOUT_SKIP_HEADERS
        ):
            continue
        # A blank header (common above weekly value columns whose real
        # label is a merged cell one row up, e.g. "Week 1") is still a
        # candidate value column - decided row-by-row on numeric-ness.
        value_cols.append(c_idx)

    if not value_cols:
        return {}, [], {}, {}, 0, False

    # First pass: collect every row that resolves to a field id (via its
    # Symbol, or its Description text as a fallback), grouped by field id,
    # instead of dumping numbers straight into field_values -- needed so
    # the "(+)"-marked-row preference below can choose among candidates
    # for the same field before anything is finalized.
    candidates = {}  # field_id -> list of {description, values, marked}
    max_readings = 0
    is_calculated_only = False
    for row in rows[header_row_idx + 1:]:
        if symbol_col >= len(row):
            continue
        sym = row[symbol_col]
        if sym is None or not str(sym).strip():
            continue
        sym_norm = re.sub(r'[^a-z0-9]', '', str(sym).lower().strip())
        if sym_norm in CALCULATED_OUTPUT_ONLY_SYMBOLS:
            is_calculated_only = True

        description = ''
        if particulars_col is not None and particulars_col < len(row) \
                and row[particulars_col] is not None:
            description = str(row[particulars_col]).strip()

        field_id = _generic_sym_to_field(sym)
        if not field_id and description:
            # Symbol didn't resolve to anything -- try the row's own
            # Description text the same way Strategy 2 matches a header
            # (exact alias, then the _match_tag_patterns rule set, e.g.
            # recovering "O2 At AH Outlet" -> O2out from wording alone).
            field_id = _label_to_field(description)
        if not field_id or field_id in NEVER_AUTO_DETECT:
            continue

        if field_id in GENERIC_ROW_LAYOUT_COMPOSITION_FIELDS \
                and _row_is_flue_gas_context(description):
            # e.g. "N2 At AH Outlet" -- a flue-gas duct reading whose
            # symbol collides with CENPEEP's coal-Nitrogen symbol, not an
            # actual coal-composition result.
            continue

        row_values = []
        for c_idx in value_cols:
            if c_idx >= len(row):
                continue
            num = _to_num(row[c_idx])
            if num is not None:
                row_values.append(num)
        if not row_values:
            continue

        max_readings = max(max_readings, len(row_values))
        candidates.setdefault(field_id, []).append({
            'description': description or str(sym).strip(),
            'values': row_values,
            'marked': '+' in description,
        })

    field_values = {}
    field_particulars = {}
    for field_id, rows_for_field in candidates.items():
        marked_rows = [r for r in rows_for_field if r['marked']]
        chosen_rows = marked_rows if marked_rows else rows_for_field
        for r in chosen_rows:
            field_values.setdefault(field_id, []).extend(r['values'])
        # "Detected From" label: sheet name + the actual Description
        # text of whichever row(s) were used, not the bare Symbol -- so
        # "where did this come from" is answerable without reopening the
        # workbook (a bare "C" or "N2" symbol is otherwise meaningless
        # out of context).
        seen_desc = []
        for r in chosen_rows:
            if r['description'] not in seen_desc:
                seen_desc.append(r['description'])
        label = ' + '.join(seen_desc)
        if sheet_name:
            label = f"{sheet_name}: {label}"
        field_particulars[field_id] = label

    extracted, raw_rows, summary = _finalize_field_values(field_values)
    col_meta = {
        fid: {
            'fieldId': fid,
            'header': field_particulars.get(fid, fid),
            'source': 'rule',
            'confidence': 1.0,
        }
        for fid in extracted
    }
    return extracted, raw_rows, summary, col_meta, max_readings, is_calculated_only


# ─── Strategy 4: Plain label/value form layout (no header row at all) ─────────
# The reference "BEE-2 (Efficiency-Indirect).xlsx" sheet (and, in general, any
# hand-built single-boiler "form" sheet, as opposed to a multi-column
# log/DCS-export table) has NO anchor/header row for Strategies 1-3 to find:
# col0 is a plain-English label ("O2 in Flue Gas"), col1 is the numeric value,
# and an optional col2 is a unit string ("%") - one parameter per row, top to
# bottom, starting from row 1. There's no "Symbol"/"Particulars"/"UOM" header
# cell anywhere (Strategy 3 needs one), and it's not a table of columns with
# a header row above data rows (Strategy 2's shape) - so every existing
# strategy correctly finds nothing on it, and uploads silently detected 0
# fields.
#
# This only activates once Strategies 1-3 have all found nothing (see
# _parse_sheet_rows / _parse_sheet_chunked), so it can't change how any
# already-working sheet is parsed.
#
# Deliberately uses EXACT alias/symbol matching only (_sym_to_field +
# LABEL_ALIASES) - NOT the fuzzy _match_tag_patterns()/ML fallback used for
# short column-header text. Row labels here are full sentences, and the
# reference sheet's own "Solution" section restates several of these same
# words in unrelated derived-value rows, e.g. "Heat loss due to unburnt in
# fly ash" (a kcal/kg figure) or "% Heat loss due to unburnts in bottom ash
# (L8)" (a completely different % figure). _match_tag_patterns' loose
# {'fly','ash'} <= tokens / {'bottom','ash'} <= tokens subset check - fine
# for terse column headers - would false-positive-match both of those onto
# Cfa/Cba and corrupt their averaged value with an unrelated number. Exact
# matching only fires on the sheet's actual parameter-label rows (which are
# also duplicated verbatim in the Solution section for a couple of fields -
# harmless, since those duplicates repeat the same value and just average
# in cleanly) and skips every derived/step-by-step row safely.
LABEL_VALUE_LAYOUT_MIN_FIELDS = 5


def _label_to_field_exact(label):
    """
    Same as _label_to_field() minus the fuzzy _match_tag_patterns() fallback
    - see the "Strategy 4" note above for why that fallback isn't safe to
    use against full-sentence row labels.
    """
    norm = re.sub(r'[^a-z0-9 ]', '', str(label).lower().strip())
    fid = _sym_to_field(label.strip())
    if fid:
        return fid
    return LABEL_ALIASES.get(norm)


def _parse_label_value_layout(rows):
    """
    Strategy 4 - see module note above. Every row is a candidate: if col0's
    text exactly matches a known field label/symbol, the first numeric cell
    elsewhere in that row is taken as its value (any non-numeric cell, e.g.
    a units string or a formula-description string, is skipped over).
    Multiple rows resolving to the same field id are averaged (same as
    every other strategy) - which is what makes the reference sheet's own
    "Solution" section safely restating a couple of these labels verbatim
    (same value) a no-op rather than a conflict.
    Requires at least LABEL_VALUE_LAYOUT_MIN_FIELDS distinct fields matched
    before "activating", so a sheet that only incidentally has one label-
    like row (e.g. a stray "Load:" cell in an unrelated table) isn't
    mistaken for this layout.
    Returns (extracted_dict, raw_rows_list, sheet_summary, col_meta,
             matched_row_count).
    """
    field_values = {}
    field_label_text = {}
    matched_rows = 0
    for row in rows:
        if len(row) < 2:
            continue
        label = row[0]
        if not isinstance(label, str) or not label.strip():
            continue
        field_id = _label_to_field_exact(label)
        if not field_id or field_id in NEVER_AUTO_DETECT:
            continue
        num = None
        for cell in row[1:]:
            num = _to_num(cell)
            if num is not None:
                break
        if num is None:
            continue
        field_values.setdefault(field_id, []).append(num)
        field_label_text.setdefault(field_id, label.strip())
        matched_rows += 1

    if matched_rows < LABEL_VALUE_LAYOUT_MIN_FIELDS:
        return {}, [], {}, {}, 0

    extracted, raw_rows, summary = _finalize_field_values(field_values)
    col_meta = {
        fid: {
            'fieldId': fid,
            'header': field_label_text.get(fid, fid),
            'source': 'rule',
            'confidence': 1.0,
        }
        for fid in extracted
    }
    return extracted, raw_rows, summary, col_meta, matched_rows


# ─── Per-sheet parser (tries all strategies) ───────────────────────────────────
def _parse_sheet_rows(rows, sheet_name, use_ml=True, highlight_map=None):
    """
    Tries CenPeep column layout first, then raw tabular layout (ML-augmented),
    then the generic labeled-row layout (Strategy 3 - only reached if both
    of the above find nothing, so it can never change how an already-working
    sheet is parsed).
    Returns a dict with keys: extracted, rawRows, strategy, summary, columns.
    """
    # Strategy 1: standard CENPEEP layout
    ext1, raw1 = _parse_cenpeep_layout(rows)
    if len(ext1) >= 5:
        return {
            'sheetName': sheet_name,
            'strategy': 'cenpeep_column',
            'extracted': ext1,
            'rawRows': raw1,
            'summary': {},
            'columns': {},
            'dataRowCount': len(rows),
            'unmatchedHighlighted': [],
            'datedRows': [],
        }

    # Strategy 3: generic labeled-row layout (position-independent Symbol
    # column, possibly several reading columns per row). Tried BEFORE
    # Strategy 2 deliberately: this layout still has a header row inside
    # Strategy 2's own header-scan window, and that header row's cell text
    # ("Parameter", "Symbol", "Unit", "Data Collection Method", a repeated
    # period label like "Boiler 1") can look just similarity-enough to real
    # CENPEEP field phrasing that Strategy 2's ML fallback grabs a few of
    # those columns with low confidence and "succeeds" with wrong values —
    # which would stop Strategy 3 from ever being tried. Strategy 3 only
    # activates at all when it finds a literal "Symbol" header cell, which
    # a genuine Strategy-2 (date/tag log) sheet essentially never has, so
    # this reordering doesn't change anything for sheets that were already
    # being parsed correctly by Strategy 2.
    ext3, raw3, summary3, col_meta3, data_row_count3, calc_only3 = _parse_generic_row_layout(
        rows, sheet_name=sheet_name
    )
    if ext3:
        return {
            'sheetName': sheet_name,
            # A calculated/reference sheet (see CALCULATED_OUTPUT_ONLY_SYMBOLS)
            # gets its own strategy tag so the merge logic in
            # _extract_from_workbook can rank it LAST instead of grouping it
            # with genuine Strategy-3 input sheets.
            'strategy': 'generic_row_layout_calculated' if calc_only3 else 'generic_row_layout',
            'extracted': ext3,
            'rawRows': raw3,
            'summary': summary3,
            'columns': col_meta3,
            'dataRowCount': data_row_count3,
            'unmatchedHighlighted': [],
            'datedRows': [],
        }

    # Strategy 2: raw tabular, with ML fallback for unrecognized headers
    ext2, raw2, summary, col_meta, data_row_count, unmatched_hi, dated_rows2 = _parse_raw_layout(
        rows, use_ml=use_ml, highlight_map=highlight_map
    )
    if ext2:
        ml_used = any(c['source'] in ('ml', 'ml_highlighted') for c in col_meta.values())
        return {
            'sheetName': sheet_name,
            'strategy': 'raw_tabular_ml' if ml_used else 'raw_tabular',
            'extracted': ext2,
            'rawRows': raw2,
            'summary': summary,
            'columns': col_meta,
            'dataRowCount': data_row_count,
            'unmatchedHighlighted': unmatched_hi,
            'datedRows': dated_rows2,
        }

    # Strategy 4: plain label/value form layout, no header row at all -
    # only reached once every other strategy has found nothing.
    ext4, raw4, summary4, col_meta4, matched_rows4 = _parse_label_value_layout(rows)
    if ext4:
        return {
            'sheetName': sheet_name,
            'strategy': 'label_value_layout',
            'extracted': ext4,
            'rawRows': raw4,
            'summary': summary4,
            'columns': col_meta4,
            'dataRowCount': matched_rows4,
            'unmatchedHighlighted': unmatched_hi,
            'datedRows': [],
        }

    return {
        'sheetName': sheet_name,
        'strategy': 'unrecognized',
        'extracted': {},
        'rawRows': [],
        'summary': {},
        'columns': {},
        'dataRowCount': 0,
        'unmatchedHighlighted': unmatched_hi,
        'datedRows': [],
    }


# ─── Workbook reader (chunked / streaming) ────────────────────────────────────
def _iter_sheet_rows_streamed(ws, ext_xls=False, xlrd_sheet=None):
    """
    Yields rows one at a time from a worksheet without materializing the
    whole sheet in memory. Works for openpyxl read_only worksheets,
    xlrd sheets (legacy .xls fallback), and calamine sheets.
    """
    if ext_xls:
        for r in range(xlrd_sheet.nrows):
            yield [xlrd_sheet.cell_value(r, c) for c in range(xlrd_sheet.ncols)]
    elif HAS_CALAMINE and isinstance(ws, python_calamine.CalamineSheet):
        for row in ws.iter_rows():
            yield row
    else:
        for row in ws.iter_rows(values_only=True):
            yield list(row)


def _parse_sheet_chunked(row_iter, sheet_name, use_ml=True, highlight_map=None):
    """
    Chunked version of sheet parsing for large sheets: reads CHUNK_ROWS rows
    at a time, identifies the header row from the first chunk, maps columns
    to fields once, then streams remaining chunks through the field-value
    accumulator and discards each chunk immediately after — so memory stays
    bounded by chunk size, not total sheet size.

    `highlight_map`, if given, is {rowIdx: {colIdx, ...}} for this sheet
    (see _scan_header_highlights) — applied right after the header row is
    identified, before any data rows are accumulated, so highlighted
    columns get the same conflict-resolution/recovery treatment as the
    non-chunked path (see _apply_highlight_signal).

    Falls back cleanly to "unrecognized" if no header / no fields found.
    Returns the same shape as _parse_sheet_rows().
    """
    chunk = []
    header_row_idx = None
    headers = None
    col_map = col_source = col_confidence = None
    field_values = {}
    cenpeep_check_rows = []  # first rows, used to check for CenPeep column layout
    row_count = 0
    date_col_idx = None
    data_row_count = 0
    unmatched_hi = []
    dated_rows = []

    for row in row_iter:
        row_count += 1
        cenpeep_check_rows_cap = 200  # CenPeep layout is always near the top
        if len(cenpeep_check_rows) < cenpeep_check_rows_cap:
            cenpeep_check_rows.append(row)

        if headers is None:
            # Still hunting for the header row in the first few rows
            chunk.append(row)
            if len(chunk) >= HEADER_SCAN_ROWS:
                idx = _find_header_row(chunk, use_ml=use_ml)
                if idx is not None:
                    header_row_idx = idx
                    raw_headers = chunk[header_row_idx]
                    headers = _stitch_stacked_headers(
                        chunk[max(0, header_row_idx - STACKED_HEADER_LOOKUP_ROWS):header_row_idx],
                        raw_headers,
                    )
                    col_map, col_source, col_confidence = _map_columns_to_fields(
                        headers, use_ml=use_ml, raw_headers=raw_headers
                    )
                    highlighted_cols = (highlight_map or {}).get(header_row_idx, set())
                    col_map, col_source, col_confidence, _, unmatched_hi = (
                        _apply_highlight_signal(col_map, col_source, col_confidence,
                                                 headers, highlighted_cols, use_ml=use_ml)
                    )
                    field_values = {fid: [] for fid in set(col_map.values())}
                    date_col_idx = _find_date_col_idx(headers)
                    # Process any data rows already buffered after the header
                    for data_row in chunk[header_row_idx + 1:]:
                        _accumulate_row(data_row, col_map, field_values,
                                         date_col_idx=date_col_idx, dated_rows=dated_rows)
                        if _row_has_data(data_row, date_col_idx, col_map):
                            data_row_count += 1
                    chunk = []
                elif len(chunk) > HEADER_SCAN_ROWS * 4:
                    # Header never found in a reasonable window — give up
                    # gracefully rather than buffering the whole sheet.
                    break
            continue

        # Header already known — accumulate this row directly, no buffering
        _accumulate_row(row, col_map, field_values,
                         date_col_idx=date_col_idx, dated_rows=dated_rows)
        if _row_has_data(row, date_col_idx, col_map):
            data_row_count += 1

    # First, check whether this is actually a strict CenPeep column-layout
    # sheet (Particulars/UOM/Symbol/Formula/Value) — that strategy wins if
    # it finds enough fields, same priority as the non-chunked path.
    ext1, raw1 = _parse_cenpeep_layout(cenpeep_check_rows)
    if len(ext1) >= 5:
        return {
            'sheetName': sheet_name,
            'strategy': 'cenpeep_column',
            'extracted': ext1,
            'rawRows': raw1,
            'summary': {},
            'columns': {},
            'rowsScanned': row_count,
            'dataRowCount': row_count,
            'unmatchedHighlighted': [],
            'datedRows': [],
        }

    # Strategy 3 fallback: generic labeled-row layout, checked against the
    # same buffered leading rows used for the CenPeep-layout check above.
    # Tried before falling back on whatever the inline Strategy-2 header
    # hunt above produced, for the same reason as in _parse_sheet_rows:
    # this layout's header row (with cell text like "Parameter", "Symbol",
    # "Data Collection Method") sits inside Strategy 2's own header-scan
    # window and can fool its ML fallback into a low-confidence, wrong
    # match before Strategy 3 ever gets a look. Strategy 3 only activates
    # when it finds a literal "Symbol" header cell, so this can't change
    # the outcome for a sheet that doesn't have this row-per-parameter
    # shape. This layout is also always a small parameter table in
    # practice (never the giant hourly-log sheets chunking exists for), so
    # checking it against cenpeep_check_rows (capped at 200 rows) rather
    # than the full streamed sheet is safe.
    ext3, raw3, summary3, col_meta3, data_row_count3, calc_only3 = _parse_generic_row_layout(
        cenpeep_check_rows, sheet_name=sheet_name
    )
    if ext3:
        return {
            'sheetName': sheet_name,
            'strategy': 'generic_row_layout_calculated' if calc_only3 else 'generic_row_layout',
            'extracted': ext3,
            'rawRows': raw3,
            'summary': summary3,
            'columns': col_meta3,
            'rowsScanned': row_count,
            'dataRowCount': data_row_count3,
            'unmatchedHighlighted': [],
            'datedRows': [],
        }

    if not col_map:
        # Strategy 4: plain label/value form layout, no header row at all -
        # checked against the same buffered leading rows as Strategies 1/3
        # above (see _parse_sheet_rows for why this shape is always a small
        # parameter table, never one of the giant sheets chunking exists
        # for, so checking only the buffered rows is safe).
        ext4, raw4, summary4, col_meta4, matched_rows4 = _parse_label_value_layout(cenpeep_check_rows)
        if ext4:
            return {
                'sheetName': sheet_name,
                'strategy': 'label_value_layout',
                'extracted': ext4,
                'rawRows': raw4,
                'summary': summary4,
                'columns': col_meta4,
                'rowsScanned': row_count,
                'dataRowCount': matched_rows4,
                'unmatchedHighlighted': unmatched_hi,
                'datedRows': [],
            }
        return {
            'sheetName': sheet_name,
            'strategy': 'unrecognized',
            'extracted': {},
            'rawRows': [],
            'summary': {},
            'columns': {},
            'rowsScanned': row_count,
            'dataRowCount': 0,
            'unmatchedHighlighted': unmatched_hi,
            'datedRows': [],
        }

    extracted, raw_rows, summary = _finalize_field_values(field_values)
    ml_used = any(col_source.get(i) in ('ml', 'ml_highlighted') for i in col_map)
    highlighted_cols_final = (highlight_map or {}).get(header_row_idx, set())
    col_meta = {
        col_idx: {
            'fieldId': fid,
            'header': str(headers[col_idx]),
            'source': col_source[col_idx],
            'confidence': col_confidence[col_idx],
            'highlighted': col_idx in highlighted_cols_final,
        }
        for col_idx, fid in col_map.items()
    }

    return {
        'sheetName': sheet_name,
        'strategy': 'raw_tabular_ml_chunked' if ml_used else 'raw_tabular_chunked',
        'extracted': extracted,
        'rawRows': raw_rows,
        'summary': summary,
        'columns': col_meta,
        'rowsScanned': row_count,
        'dataRowCount': data_row_count,
        'unmatchedHighlighted': unmatched_hi,
        'datedRows': dated_rows,
    }


def _accumulate_row(row, col_map, field_values, date_col_idx=None, dated_rows=None):
    """
    Pull numeric values for mapped columns out of one data row. If
    `dated_rows` is given (a list), also append this row's dated snapshot
    to it — same shape/purpose as in _parse_raw_layout, capped at
    MAX_DATED_ROWS_PER_SHEET — so the chunked/streamed path (the one large,
    real hourly-log workbooks actually go through) supports the date-range
    feature too, not just the small-sheet path.
    """
    row_vals = {}
    row_sums = {}
    row_counts = {}
    for col_idx, fid in col_map.items():
        val = row[col_idx] if col_idx < len(row) else None
        num = _to_num(val)
        if num is not None and not _reading_in_physical_range(fid, num):
            num = None
        if num is not None:
            if fid not in MULTI_COLUMN_SUM_FIELDS:
                field_values.setdefault(fid, []).append(num)
            # A field can have MORE THAN ONE column on the same row (see
            # MULTI_COLUMN_AVERAGE_FIELDS, e.g. O2in's Left/Right duct O2
            # columns) — average them together for this row's dated
            # snapshot instead of letting the second column's value
            # silently overwrite the first. Without this, the whole-file
            # value (built from field_values, which correctly pools every
            # column's readings) stayed right, but the per-date average a
            # person sees after picking a date range on the calculator
            # silently used only one of the two columns for every row.
            # MULTI_COLUMN_SUM_FIELDS (e.g. Fpa/Fsa — "Total" duct flow,
            # see the MULTI_COLUMN_SUM_FIELDS comment) instead want that
            # row's columns SUMMED, not averaged.
            row_sums[fid] = row_sums.get(fid, 0.0) + num
            row_counts[fid] = row_counts.get(fid, 0) + 1
    for fid, total in row_sums.items():
        if fid in MULTI_COLUMN_SUM_FIELDS:
            row_vals[fid] = total
            field_values.setdefault(fid, []).append(total)
        else:
            row_vals[fid] = total / row_counts[fid]
    if dated_rows is not None and row_vals and len(dated_rows) < MAX_DATED_ROWS_PER_SHEET:
        row_date = None
        if date_col_idx is not None and date_col_idx < len(row):
            row_date = _parse_date_cell(row[date_col_idx])
        dated_rows.append({
            'date': row_date.isoformat() if row_date else None,
            'values': row_vals,
        })


def _is_readable_worksheet(ws):
    """
    True if `ws` is a normal cell-grid worksheet (openpyxl Worksheet or
    ReadOnlyWorksheet) rather than a Chartsheet/Dialogsheet or other
    non-worksheet type with no cell grid to read.

    Checked by capability (has iter_rows) rather than isinstance, since
    read_only=True workbooks use ReadOnlyWorksheet, not the normal
    Worksheet class.
    """
    return hasattr(ws, 'iter_rows')


def _sheet_row_estimate(ws):
    """Best-effort row count for an openpyxl worksheet (read_only safe)."""
    try:
        return ws.max_row or 0
    except Exception:
        return 0


def _parse_with_calamine(file_bytes, use_ml=True, highlight_map=None):
    """
    calamine (Rust) reads the workbook's shared strings/values directly
    without building openpyxl's Python style-object graph. This matters
    a lot in practice: real plant/historian exports often carry tens of
    thousands of near-duplicate per-cell styles, and openpyxl's style
    resolution during load_workbook() can dominate parse time — e.g. a
    ~5MB sheet with ~46,000 named styles took ~26s to even open with
    openpyxl vs ~0.1s with calamine. This is what makes parsing
    100MB-class uploads practical within a normal request timeout.

    NOTE: calamine can raise pyo3_runtime.PanicException (a BaseException,
    not an Exception) if its Rust parser trips on an unusual/malformed
    file. Callers must be prepared to catch BaseException around this
    call, not just Exception — see parse_workbook().
    """
    sheet_results = []
    wb = python_calamine.CalamineWorkbook.from_filelike(io.BytesIO(file_bytes))
    for meta in wb.sheets_metadata:
        name = meta.name
        # Skip chart sheets / dialog sheets / macro sheets — they carry
        # no tabular cell data. Checked via explicit sheet-type metadata
        # (robust) rather than trying to read them and catching the
        # resulting AttributeError.
        if meta.typ != python_calamine.SheetTypeEnum.WorkSheet:
            sheet_results.append({
                'sheetName': name,
                'strategy': 'skipped_non_worksheet',
                'extracted': {},
                'rawRows': [],
                'summary': {},
                'columns': {},
                'rowsScanned': 0,
            })
            continue
        ws = wb.get_sheet_by_name(name)
        est_rows = ws.height or 0
        row_iter = _iter_sheet_rows_streamed(ws)
        sheet_hi_map = (highlight_map or {}).get(name, {})
        if est_rows > LARGE_SHEET_ROW_THRESHOLD:
            result = _parse_sheet_chunked(row_iter, name, use_ml=use_ml, highlight_map=sheet_hi_map)
        else:
            rows = list(row_iter)
            result = _parse_sheet_rows(rows, name, use_ml=use_ml, highlight_map=sheet_hi_map)
        sheet_results.append(result)
    return sheet_results


def _parse_with_xlrd(file_bytes, use_ml=True, highlight_map=None):
    """Legacy .xls reader (calamine/openpyxl don't handle old binary .xls)."""
    sheet_results = []
    wb = xlrd.open_workbook(file_contents=file_bytes)
    for name in wb.sheet_names():
        ws = wb.sheet_by_name(name)
        row_iter = _iter_sheet_rows_streamed(None, ext_xls=True, xlrd_sheet=ws)
        sheet_hi_map = (highlight_map or {}).get(name, {})
        if ws.nrows > LARGE_SHEET_ROW_THRESHOLD:
            result = _parse_sheet_chunked(row_iter, name, use_ml=use_ml, highlight_map=sheet_hi_map)
        else:
            rows = list(row_iter)
            result = _parse_sheet_rows(rows, name, use_ml=use_ml, highlight_map=sheet_hi_map)
        sheet_results.append(result)
    return sheet_results


def _parse_with_openpyxl(file_bytes, use_ml=True, highlight_map=None):
    """
    Pure-Python fallback reader. Slower than calamine on files with heavy
    style bloat, but more lenient — used both when calamine isn't installed
    and when calamine fails to read a particular file (see parse_workbook()).
    """
    sheet_results = []
    # read_only=True streams the worksheet instead of materializing
    # the whole workbook as Cell objects — this is the key change that
    # lets large files (10MB+, wide sheets, 1000+ rows) parse without
    # blowing up memory the way the original full-load approach did.
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        # Chartsheets (and other non-worksheet sheet types, e.g. dialogsheets)
        # have no cell grid and no iter_rows() — attempting to read them the
        # same way as a normal worksheet raises
        # "'Chartsheet' object has no attribute 'iter_rows'". Skip them
        # instead of crashing the whole upload. Checked by capability
        # (has iter_rows) rather than isinstance, since read_only=True
        # workbooks use ReadOnlyWorksheet, not the normal Worksheet class.
        if not _is_readable_worksheet(ws):
            sheet_results.append({
                'sheetName': name,
                'strategy': 'skipped_non_worksheet',
                'extracted': {},
                'rawRows': [],
                'summary': {},
                'columns': {},
                'rowsScanned': 0,
            })
            continue
        est_rows = _sheet_row_estimate(ws)
        row_iter = _iter_sheet_rows_streamed(ws)
        sheet_hi_map = (highlight_map or {}).get(name, {})
        if est_rows > LARGE_SHEET_ROW_THRESHOLD:
            result = _parse_sheet_chunked(row_iter, name, use_ml=use_ml, highlight_map=sheet_hi_map)
        else:
            rows = list(row_iter)
            result = _parse_sheet_rows(rows, name, use_ml=use_ml, highlight_map=sheet_hi_map)
        sheet_results.append(result)
    wb.close()
    return sheet_results


# ─── Main parse entry-point ───────────────────────────────────────────────────
def parse_workbook(file_bytes, filename, use_ml=True):
    """
    Parse all sheets, automatically choosing chunked streaming for large
    sheets (row count above LARGE_SHEET_ROW_THRESHOLD) and the simpler
    in-memory path for small ones. Returns:
      {
        sheetResults: [ { sheetName, strategy, extracted, rawRows, summary, columns }, … ],
        extracted:    { merged field dict — CenPeep sheet wins },
        primarySheet: str,
        totalFields:  int,
        parseTimeMs:  float,
      }
    """
    t_start = time.time()
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    sheet_results = []

    # Highlight-based detection is intentionally DISABLED — field detection
    # must come from header-text matching (rule + ML) alone, never from a
    # cell's fill color. Real-world workbooks routinely apply a themed
    # header-row style (banded/table header formatting) across an ENTIRE
    # header row for purely cosmetic reasons; _is_highlighted_fill() can't
    # tell that apart from an engineer deliberately marking one column, so
    # every column on such a sheet was being treated as "human-flagged" —
    # e.g. this caused a sheet with both an "IM %" and a "T.M. %" column to
    # have BOTH pulled into the Moisture field (their headers joined as
    # "IM % + T.M. %") instead of just the correct T.M. one, purely because
    # the sheet's header row happened to use a themed cell style.
    # _scan_header_highlights / _apply_highlight_signal are left in place
    # below (harmless, unused) rather than deleted, in case this needs a
    # narrower, more reliable re-enable later — but they are never invoked
    # from here, so highlight_map is always empty and every "highlighted"
    # code path downstream is a permanent no-op.
    highlight_map = {}

    use_calamine = HAS_CALAMINE
    if use_calamine:
        try:
            sheet_results = _parse_with_calamine(file_bytes, use_ml=use_ml, highlight_map=highlight_map)
        except BaseException as e:
            # calamine is a compiled Rust extension. When its parser hits
            # something it can't handle it doesn't raise a normal Python
            # Exception — it raises pyo3_runtime.PanicException, which is
            # deliberately made to inherit from BaseException (not
            # Exception) specifically so a bare `except Exception` anywhere
            # upstream (e.g. this route's error handler) will NOT catch it.
            # Left alone, that blows straight through Flask's request
            # handling and the client gets a dead connection with no
            # response at all, instead of a normal error. Let real
            # interpreter-shutdown signals through; treat everything else
            # (including PanicException) as "this file broke calamine" and
            # fall back to the openpyxl/xlrd reader instead of taking the
            # whole upload down.
            if isinstance(e, (KeyboardInterrupt, SystemExit)):
                raise
            use_calamine = False
            sheet_results = []

    if not use_calamine:
        if ext == 'xls':
            if not HAS_XLRD:
                raise RuntimeError('xlrd not installed; cannot read .xls files')
            sheet_results = _parse_with_xlrd(file_bytes, use_ml=use_ml, highlight_map=highlight_map)
        else:
            if not HAS_OPENPYXL:
                raise RuntimeError('Neither python-calamine nor openpyxl is installed')
            sheet_results = _parse_with_openpyxl(file_bytes, use_ml=use_ml, highlight_map=highlight_map)

    # Merge: pick ONE "best" sheet among the generic (non-CenPeep) sheets —
    # whichever has the most real (non-blank) date rows — and take ALL of
    # its extracted fields as-is. A field's average/sum must never be mixed
    # across sheets; it always comes wholly from one sheet's rows.
    #
    # Only fields the best sheet doesn't have at all are backfilled from the
    # next-best sheet (2nd-most date rows), then the next, and so on. This
    # is deliberately a SHEET-level choice, not a per-field one — e.g. an
    # "Hourly" log sheet (many date rows) should be used wholesale over a
    # "Day Avg" summary sheet (few rows, itself just an average of the
    # hourly sheet), rather than comparing row counts field by field.
    #
    # The CenPeep column-layout sheet, if present, still wins over both
    # (it's a distinct, authoritative single-value layout, not a log).
    #
    # Raw data-row COUNT alone isn't always enough to find the real master
    # log, though: a plant workbook often also carries a filtered/derived
    # subset of the same columns for one specific analysis slice (e.g. a
    # "LOAD BAND" sheet covering only one load range), and that filtered
    # sheet can end up with a row count that ties or even edges out the
    # actual full-period sheet depending on how blanks/merged cells land.
    # "OPN PARA" ("Operation Parameters") is the standard real-plant sheet
    # name for the full-period master log this data comes from, so a sheet
    # matching that naming convention is preferred as the primary raw
    # source ahead of plain row-count ranking. This is purely a tiebreak/
    # preference signal, not a requirement -- workbooks without a sheet
    # named this way fall back to row-count ranking exactly as before.
    merged_extracted = {}
    merged_field_source = {}   # field_id -> (sheetName, dataRowCount) chosen
    merged_field_detail = {}   # field_id -> {sheet, header, source, confidence}
    cenpeep_result = None
    generic_row_layout_sheets = []
    calculated_reference_sheets = []
    generic_sheets = []
    for sr in sheet_results:
        if 'cenpeep' in sr['sheetName'].lower():
            cenpeep_result = sr
            continue
        if sr.get('strategy') == 'generic_row_layout_calculated':
            # A Strategy-3 sheet whose Symbol column contains
            # CALCULATED_OUTPUT_ONLY_SYMBOLS (dry-gas loss, unburnt-carbon
            # loss, etc) -- e.g. a "BOILER EFFICIENCY CALCULATION" sheet
            # the user keeps purely as a computed reference/cross-check,
            # not a raw input source. It legitimately re-lists real input
            # parameters (O2in, Tgo, GCV, Cba, Cfa, ...) alongside its
            # computed losses, which is exactly why it must NOT get the
            # same "always wins" priority as a genuine Strategy-3 input
            # sheet below -- otherwise it silently overrides the real raw
            # log sheet's values with its own derived/summarized numbers.
            # Kept only as an absolute last-resort backfill for fields no
            # other sheet found at all.
            calculated_reference_sheets.append(sr)
            continue
        if sr.get('strategy') == 'generic_row_layout':
            # Same reasoning as the cenpeep_result carve-out above: a sheet
            # that matched Strategy 3 was identified via a strong,
            # unambiguous signal (a literal "Symbol" column header on a
            # dedicated row-per-parameter table) rather than fuzzy
            # header-text similarity, so it's a deliberate efficiency-input
            # sheet, not a raw log/design dump. It should outrank ordinary
            # raw_tabular sheets the same way the named CenPeep sheet does
            # - otherwise a bigger but noisier log sheet (more date rows,
            # but ML-matched with lower confidence) can win the ranking
            # below and silently override a correct, high-confidence value
            # with a wrong one.
            generic_row_layout_sheets.append(sr)
            continue
        generic_sheets.append(sr)

    # Rank generic sheets by (a) whether the sheet follows the real-plant
    # "OPN PARA" master-log naming convention, then (b) date-row count,
    # most rows first. (a) is a preference tiebreak, not a hard filter --
    # see the comment above.
    ranked_sheets = sorted(
        generic_sheets,
        key=lambda sr: (_is_opn_para_sheet(sr['sheetName']), sr.get('dataRowCount', 0)),
        reverse=True,
    )
    # generic_row_layout sheets are tried first (see comment above), each
    # one ranked by how many fields it found, before falling back to the
    # ordinary date-row-count ranked sheets for anything still missing.
    # Calculated/reference sheets (see comment above) go dead last -- after
    # every other sheet, including plain raw-tabular ones -- so they only
    # ever backfill a field nothing else on the workbook could find.
    ranked_sheets = sorted(
        generic_row_layout_sheets, key=lambda sr: len(sr.get('extracted', {})), reverse=True
    ) + ranked_sheets + sorted(
        calculated_reference_sheets, key=lambda sr: len(sr.get('extracted', {})), reverse=True
    )

    best_generic_sheet = None
    for sr in ranked_sheets:
        if not sr['extracted']:
            continue
        sr_rows = sr.get('dataRowCount', 0)
        if best_generic_sheet is None:
            best_generic_sheet = sr
        field_details = _sheet_field_details(sr)
        # Take every field this sheet has, but only if the field hasn't
        # already been filled by a higher-ranked (more date rows) sheet.
        for fid, val in sr['extracted'].items():
            if fid in merged_extracted:
                continue
            merged_extracted[fid] = val
            merged_field_source[fid] = (sr['sheetName'], sr_rows)
            detail = field_details.get(fid, {})
            merged_field_detail[fid] = {
                'sheet': sr['sheetName'],
                'label': FIELD_LABELS.get(fid, fid),
                'header': detail.get('header'),
                'source': detail.get('source', 'rule'),
                'confidence': detail.get('confidence', 1.0),
            }

    # Cross-sheet override for MULTI_COLUMN_SUM_FIELDS (Fpa/Fsa): the
    # ranking loop above is deliberately a whole-SHEET choice (see comment
    # above), so a field only gets backfilled from a lower-ranked sheet
    # when the higher-ranked sheet doesn't have it AT ALL — even a weak,
    # ML-matched per-mill SUM on the top-ranked raw log sheet blocks a
    # much stronger, deterministically rule-matched precomputed "Total
    # PA/SA Flow" column that happens to live on a different (lower
    # row-count) sheet. Real plant workbooks commonly split this way: the
    # raw per-mill columns live on the big hourly log, but the engineer's
    # own precomputed total lives on a separate period-average sheet. For
    # just these SUM fields, a rule-matched total on ANY sheet is trusted
    # over an ML-matched per-mill sum on another sheet — the same "rule
    # beats ML" precedence _dedupe_columns_per_field already applies
    # WITHIN one sheet's columns, extended across sheets here.
    for fid in MULTI_COLUMN_SUM_FIELDS:
        current_detail = merged_field_detail.get(fid)
        if not current_detail or current_detail.get('source') != 'ml':
            continue
        for sr in ranked_sheets:
            if sr['sheetName'] == current_detail.get('sheet'):
                continue
            if fid not in sr.get('extracted', {}):
                continue
            alt_detail = _sheet_field_details(sr).get(fid, {})
            if alt_detail.get('source') != 'rule':
                continue
            merged_extracted[fid] = sr['extracted'][fid]
            merged_field_source[fid] = (sr['sheetName'], sr.get('dataRowCount', 0))
            merged_field_detail[fid] = {
                'sheet': sr['sheetName'],
                'label': FIELD_LABELS.get(fid, fid),
                'header': alt_detail.get('header'),
                'source': 'rule',
                'confidence': alt_detail.get('confidence', 1.0),
            }
            break

    # Fallback primary sheet: the best-ranked generic sheet if one produced
    # fields; otherwise whichever sheet produced the most fields at all, so
    # a chart sheet / empty cover sheet can't "win" despite contributing
    # nothing.
    if best_generic_sheet is not None:
        primary_sheet = best_generic_sheet['sheetName']
    elif sheet_results:
        primary_sheet = max(sheet_results, key=lambda sr: len(sr['extracted']))['sheetName']
    else:
        primary_sheet = ''
    if cenpeep_result:
        cenpeep_details = _sheet_field_details(cenpeep_result)
        for fid in cenpeep_result['extracted']:
            merged_field_source[fid] = (cenpeep_result['sheetName'], None)
            merged_field_detail[fid] = {
                'sheet': cenpeep_result['sheetName'],
                'label': FIELD_LABELS.get(fid, fid),
                'header': None,
                'source': 'cenpeep_column',
                'confidence': cenpeep_details.get(fid, {}).get('confidence', 1.0),
            }
        merged_extracted.update(cenpeep_result['extracted'])
        primary_sheet = cenpeep_result['sheetName']

    # Fallback: Primary Air to APH Temp In (Tpai) very often has no column
    # of its own on real plant sheets — only the Secondary Air In (Tsai)
    # reading is logged. Physically, ambient air entering the APH is drawn
    # from the same source for both the primary- and secondary-air ducts,
    # so when Tpai genuinely wasn't found anywhere but Tsai WAS, default
    # Tpai to the same value as Tsai instead of leaving it blank/missing.
    # This only fires when Tpai is absent from every sheet — it never
    # overrides an actually-detected Tpai value.
    if 'Tpai' not in merged_extracted and 'Tsai' in merged_extracted:
        merged_extracted['Tpai'] = merged_extracted['Tsai']
        tsai_source = merged_field_source.get('Tsai')
        if tsai_source:
            merged_field_source['Tpai'] = tsai_source
        tsai_detail = merged_field_detail.get('Tsai', {})
        merged_field_detail['Tpai'] = {
            'sheet': tsai_detail.get('sheet'),
            'label': FIELD_LABELS.get('Tpai', 'Tpai'),
            'header': f"defaulted = Secondary Air In ({tsai_detail.get('header') or 'Tsai'})",
            'source': 'derived_fallback',
            'confidence': tsai_detail.get('confidence', 1.0),
        }

    missing_fields = [
        {'id': fid, 'label': FIELD_LABELS.get(fid, fid)}
        for fid in REQUIRED_FIELDS if fid not in merged_extracted
    ]

    # ─── Date-wise process support ────────────────────────────────────────
    # Per-row dated snapshots must come from EVERY sheet that actually
    # backs a merged field value — not just best_generic_sheet. Real plant
    # workbooks routinely split raw data across multiple sheets with
    # different logging cadences (e.g. an hourly "OPN PARA" operating log
    # plus a separate once-a-day "LAB" sheet for coal-quality figures, or a
    # "GUHR Direct" sheet for GCV/coal-rate). best_generic_sheet is only
    # ever ONE of those — it's whichever sheet has the most date rows
    # overall — so a field whose real source is a different, lower-row-
    # count sheet (Fixed Carbon, GCV, Moisture, Ash, Volatile Matter,
    # Unburnt Carbon in Fly/Bottom Ash, …) previously had NO per-date rows
    # at all. When a person then picked a start/end date on the
    # calculator, _averageFieldsInRange() on the frontend only had
    # best_generic_sheet's rows to filter, so those lab-only fields never
    # appeared in `avg` and silently fell back to whatever was already in
    # the form — the field's average across the sheet's ENTIRE row range
    # (every process/date block in the workbook combined), not just the
    # rows inside the selected date range. That's the "average of the
    # whole [sheet] instead of the selected date" bug.
    #
    # Fix: walk every sheet in the SAME priority order used to build
    # merged_extracted above (ranked_sheets — CenPeep-column sheets are
    # handled separately below since they're not a time series at all),
    # and for each sheet pull dated values ONLY for the fields that sheet
    # actually won in that merge (merged_field_source[fid][0] ==
    # this sheet's name). That keeps every field's date-filtered average
    # sourced from the exact same sheet as its whole-file value — a field
    # is never mixed across sheets, only re-averaged over fewer rows of
    # its own sheet — while still letting fields from DIFFERENT sheets
    # each be re-averaged over their own matching dates when a date range
    # is chosen.
    # NOTE: this stays a flat LIST of rows (one entry per original data
    # row), never a dict keyed by date. A sheet like "OPN PARA" logs many
    # rows per day (hourly readings) and every one of them must survive
    # separately so _averageFieldsInRange() on the frontend keeps summing
    # all of them for a date — collapsing same-date rows together here
    # would silently throw away all but one hourly reading per day.
    primary_dated_rows = []
    if not cenpeep_result:
        for sr in ranked_sheets:
            sheet_name = sr['sheetName']
            sheet_fields = {
                fid for fid, src in merged_field_source.items()
                if src[0] == sheet_name
            }
            if not sheet_fields:
                continue
            for row in sr.get('datedRows', []):
                row_date = row.get('date')
                if not row_date:
                    continue
                row_vals = {
                    fid: val for fid, val in row.get('values', {}).items()
                    if fid in sheet_fields
                }
                if not row_vals:
                    continue
                primary_dated_rows.append({'date': row_date, 'values': row_vals})

    # CenPeep-column layout is a single authoritative value per field, not
    # a time series, so there's no meaningful date range to slice it by —
    # matches the previous behavior for that case (primary_dated_rows stays
    # []  above when cenpeep_result is set).
    available_dates = sorted({r['date'] for r in primary_dated_rows if r.get('date')})

    # Highlighted header cells that never resolved to any CENPEEP field on
    # their sheet, even after the relaxed highlight-only retry — these are
    # exactly the columns a person marked as important that the parser
    # genuinely couldn't place, so they're worth a human glance rather than
    # being silently dropped like every other unmatched column.
    unmatched_highlighted = [
        {'sheet': sr['sheetName'], 'header': item['header']}
        for sr in sheet_results
        for item in sr.get('unmatchedHighlighted', [])
    ]

    return {
        'sheetResults': sheet_results,
        'extracted': merged_extracted,
        # Highlighted columns that still couldn't be mapped to a field —
        # surface these explicitly instead of letting them disappear into
        # the general "unrecognized column" pile.
        'unmatchedHighlighted': unmatched_highlighted,
        # Which sheet each field's final value was actually taken from —
        # use this to sanity-check that (e.g.) Load came from the Hourly
        # sheet and not the Day Avg sheet.
        'fieldSource': {fid: src[0] for fid, src in merged_field_source.items()},
        # Per-field detail on the SELECTED sheet's decision: which sheet,
        # which header text (if any), rule vs ML, and confidence — this is
        # what the "selected field + confidence" summary is built from,
        # instead of re-deriving it from every sheet's raw column list.
        'fieldDetail': merged_field_detail,
        'primarySheet': primary_sheet,
        'totalFields': len(merged_extracted),
        # Required CENPEEP input fields that were NOT found on any sheet —
        # i.e. still need to be entered manually.
        'missingFields': missing_fields,
        'parseTimeMs': round((time.time() - t_start) * 1000, 1),
        # Per-row dated snapshots ({date, values}) from the sheet that
        # backs the merged fields, and the sorted list of distinct dates
        # that have at least one — used by the frontend to (a) put a red
        # dot on the calendar for dates with real data, and (b) re-average
        # just the rows inside a chosen start/end range into a separate
        # "process" instead of the whole-file average, without another
        # upload/reparse round-trip.
        'datedRows': primary_dated_rows,
        'availableDates': available_dates,
        'dateFilteringAvailable': bool(available_dates),
    }


# ─── Field-level confidence lookup (for the "selected sheet" summary) ───────
def _sheet_field_details(sr):
    """
    Returns {fieldId: {'header': str|None, 'source': str, 'confidence': float}}
    for one sheet's result — regardless of which strategy produced it.

    'cenpeep_column' rows are an exact Particulars/Symbol/Value layout match,
    so every field from it is confidence 1.0 with no header-guessing involved.
    'raw_tabular*' sheets carry per-column metadata (rule match vs ML match,
    with the ML confidence score) in sr['columns'], keyed by column index —
    this re-keys that by field id instead, which is what a "what was picked,
    and how sure were we" summary actually needs.
    """
    if sr['strategy'] == 'cenpeep_column':
        return {
            fid: {'header': None, 'source': 'cenpeep_column', 'confidence': 1.0}
            for fid in sr['extracted']
        }
    details = {}
    for meta in sr.get('columns', {}).values():
        fid = meta['fieldId']
        if fid in details:
            # Multiple columns mapped to the same field id — this only
            # happens for MULTI_COLUMN_AVERAGE_FIELDS (e.g. O2out's
            # Left/Right APH columns, see _dedupe_columns_per_field), so
            # combine both headers into one "Detected From" string instead
            # of letting whichever column came last silently overwrite the
            # other in this fid-keyed summary.
            existing = details[fid]
            details[fid] = {
                'header': f"{existing['header']} + {meta['header']}" if existing['header'] else meta['header'],
                'source': meta['source'],
                'confidence': max(existing['confidence'] or 0, meta['confidence'] or 0),
            }
        else:
            details[fid] = {
                'header': meta['header'],
                'source': meta['source'],
                'confidence': meta['confidence'],
            }
    return details


# ─── Route ────────────────────────────────────────────────────────────────────
ALLOWED_EXTS = {'.xlsx', '.xls', '.xlsm'}


@upload_bp.route('/', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file uploaded'}), 400

    f = request.files['file']
    if not f.filename:
        return jsonify({'ok': False, 'error': 'Empty filename'}), 400

    ext = '.' + f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    if ext not in ALLOWED_EXTS:
        return jsonify({'ok': False, 'error': 'Only .xlsx / .xls / .xlsm files are accepted'}), 400

    try:
        file_bytes = f.read()
        result = parse_workbook(file_bytes, f.filename, use_ml=True)
        primary_sheet_result = next(
            (sr for sr in result['sheetResults'] if sr['sheetName'] == result['primarySheet']),
            result['sheetResults'][0] if result['sheetResults'] else None,
        )
        return jsonify({
            'ok': True,
            'filename': f.filename,
            'fileSizeMB': round(len(file_bytes) / (1024 * 1024), 2),
            **result,
            # Keep legacy fields for backward compat with existing frontend
            'sheetName': result['primarySheet'],
            'rawRows': primary_sheet_result['rawRows'] if primary_sheet_result else [],
        })
    except MemoryError:
        return jsonify({
            'ok': False,
            'error': 'File is too large to process even with chunked parsing. '
                     'Try splitting it into smaller sheets/files.',
        }), 413
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    except BaseException as e:
        # Defense-in-depth: parse_workbook() already falls back to openpyxl
        # if calamine panics (see _parse_with_calamine), but if some other
        # BaseException-derived error (e.g. a pyo3 PanicException from a
        # different native call) slips through, catch it here too rather
        # than letting it kill the request with no response at all.
        if isinstance(e, (KeyboardInterrupt, SystemExit)):
            raise
        return jsonify({'ok': False, 'error': f'Unexpected parser failure: {e}'}), 500


@upload_bp.route('/retrain', methods=['POST'])
def retrain_model():
    """
    Retrains the ML field classifier from the current contents of
    ml/training_data.py and persists it to disk. Call this after editing
    training_data.py (adding new header phrasings, fixing a mislabeled
    example, etc.) so changes take effect without restarting the server.
    """
    # Vercel's filesystem is read-only under /var/task. Retraining and saving
    # to disk is therefore disabled there. The API still works because the
    # classifier can train in memory and/or use /tmp via ml/field_classifier.py.
    if os.getenv("VERCEL", "").lower() in {"1", "true", "yes"}:
        return jsonify({
            'ok': False,
            'error': 'Retraining is disabled on Vercel because the filesystem is read-only.',
        }), 400

    try:
        from ml.field_classifier import retrain_and_save
        clf = retrain_and_save()
        return jsonify({
            'ok': True,
            'trainingExamples': len(clf.train_labels),
            'message': 'Field classifier retrained successfully.',
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500